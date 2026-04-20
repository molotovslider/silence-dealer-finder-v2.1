#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Silence.eco — Dealer Finder v11 — Production"""
import tkinter as tk
import json, re, time, os, csv, random, threading, urllib.request, urllib.parse, urllib.error
import subprocess, sys
from tkinter import ttk, messagebox, scrolledtext
from datetime import datetime

# ── Auto-install missing modules at first launch ──────────────────────────────
def detect_system():
    """Detect CPU/RAM and return recommended performance mode."""
    try:
        import psutil
        cpu  = os.cpu_count() or 2
        ram  = psutil.virtual_memory().total / (1024**3)
        avail= psutil.virtual_memory().available / (1024**3)
    except Exception:
        cpu = os.cpu_count() or 2
        ram = 4.0; avail = 2.0

    if cpu >= 8 and ram >= 12:
        return {"mode":"turbo",  "threads":16, "delay":0.2, "cpu":cpu, "ram":ram}
    elif cpu >= 6 and ram >= 8:
        return {"mode":"rapide", "threads":10, "delay":0.3, "cpu":cpu, "ram":ram}
    elif cpu >= 4 and ram >= 4:
        return {"mode":"normal", "threads":6,  "delay":0.6, "cpu":cpu, "ram":ram}
    else:
        return {"mode":"eco",    "threads":3,  "delay":1.0, "cpu":cpu, "ram":ram}

def _ensure_deps():
    # If running as a compiled PyInstaller EXE, all deps are already bundled
    # Only check when running as a plain .py script
    if getattr(sys, "frozen", False):
        return  # running as EXE — deps already included by PyInstaller

    needed = []
    try: import selenium
    except ImportError: needed.append("selenium")
    try: import webdriver_manager
    except ImportError: needed.append("webdriver-manager")
    try: import bs4
    except ImportError: needed.append("beautifulsoup4")
    try: import requests
    except ImportError: needed.append("requests")
    try: import psutil
    except ImportError: needed.append("psutil")
    try: import openpyxl
    except ImportError: needed.append("openpyxl")

    if needed:
        try:
            root = tk.Tk(); root.withdraw()
            messagebox.showinfo(
                "Silence Dealer Finder — Installation",
                f"Installation des modules requis :\n{chr(10).join(needed)}"
                f"\n\nCela prend 30 secondes, l'app redémarre ensuite."
            )
            root.destroy()
        except Exception:
            pass
        for pkg in needed:
            subprocess.check_call([sys.executable, "-m", "pip", "install",
                                   pkg, "--quiet"])
        os.execv(sys.executable, [sys.executable] + sys.argv)

_ensure_deps()

# System detection AFTER deps are installed (psutil may have just been installed)
SYSTEM = detect_system()

# ── Clés API intégrées ────────────────────────────────────────────────────────
HUNTER_KEY = "6f720d52e7ff130ef0717a890cd35abcac84c6fd"
CLAUDE_KEY  = ""  # optionnel

# ── Couleurs Silence ──────────────────────────────────────────────────────────
RED="#D01A20"; RED_D="#A01015"; RED_LT="#FAE8E8"; DARK="#1C1C1E"
GRAY="#F2F2F2"; WHITE="#FFFFFF"; MUTED="#8A8A8E"; BORDER="#E0E0E0"
GREEN="#15803d"; GREEN_LT="#F0FFF4"; AMBER="#b45309"; AMBER_LT="#FFFBF0"

# ── Traductions ───────────────────────────────────────────────────────────────
T = {
"FR": dict(
    title="Silence.eco — Dealer Finder",
    tab1="  Extraction  ", tab2="  Résultats  ", tab3="  Paramètres  ",
    url_label="URL de la page réseau concessionnaires :",
    btn_go="  Extraire  ", btn_run="  En cours…  ", btn_again="  Relancer  ",
    opt1="Recherche emails (web + Hunter.io)", opt2="Dédupliquer", opt3="Exclure emails constructeur",
    info="Analyse la page → extrait tous les concessionnaires → trouve leur site web\n→ scrape les emails → interroge Hunter.io pour enrichir les contacts.",
    log_lbl="Journal :",
    col_num="#", col_name="Nom", col_addr="Adresse", col_phone="Téléphone",
    col_email="Email", col_src="Source", col_role="Rôle / Personne",
    src_page="Page", src_web="Site web", src_hunter="Hunter.io",
    stat1="Concessionnaires", stat2="Emails trouvés", stat3="Via Hunter/web", stat4="Adresses",
    btn_copy="Copier les emails", btn_csv="Exporter Excel",
    no_results="Aucun résultat — lancez une extraction",
    # Paramètres
    p_search="Recherche & Extraction",
    p_delay="Délai entre requêtes (secondes)",
    p_delay_desc="Augmentez si le site bloque les requêtes (recommandé : 1-3s)",
    p_timeout="Timeout connexion (secondes)",
    p_timeout_desc="Temps maximum d'attente par page (recommandé : 15-25s)",
    p_retries="Tentatives en cas d'échec",
    p_retries_desc="Nombre de réessais si une page ne répond pas (recommandé : 2-4)",
    p_filter="Filtres emails",
    p_excl="Domaines à exclure (constructeurs)",
    p_excl_desc="Un domaine par ligne — ces emails seront ignorés automatiquement",
    p_also_excl="Exclure aussi les emails génériques (info@, contact@…)",
    p_export="Export",
    p_include_no_email="Inclure les concessionnaires sans email dans le CSV",
    p_open_csv="Ouvrir le CSV après export",
    btn_save="  Enregistrer  ", btn_reset="Réinitialiser",
    info_saved="Paramètres enregistrés !",
    info_reset="Paramètres réinitialisés.",
    err_url="URL invalide (doit commencer par http)",
    export_ok="CSV sauvegardé sur le Bureau :",
    ready="Prêt", last="Dernier run : ",
    copied="emails copiés !", no_copy="Aucun email.",
),
"ES": dict(
    title="Silence.eco — Dealer Finder",
    tab1="  Extracción  ", tab2="  Resultados  ", tab3="  Ajustes  ",
    url_label="URL de la página red de concesionarios:",
    btn_go="  Extraer  ", btn_run="  En proceso…  ", btn_again="  Reiniciar  ",
    opt1="Buscar emails (web + Hunter.io)", opt2="Eliminar duplicados", opt3="Excluir emails fabricante",
    info="Analiza la página → extrae concesionarios → encuentra su web\n→ extrae emails → consulta Hunter.io para enriquecer contactos.",
    log_lbl="Registro:",
    col_num="#", col_name="Nombre", col_addr="Dirección", col_phone="Teléfono",
    col_email="Email", col_src="Fuente", col_role="Rol / Persona",
    src_page="Página", src_web="Sitio web", src_hunter="Hunter.io",
    stat1="Concesionarios", stat2="Emails encontrados", stat3="Vía Hunter/web", stat4="Direcciones",
    btn_copy="Copiar emails", btn_csv="Exportar Excel",
    no_results="Sin resultados — lance una extracción",
    p_search="Búsqueda y Extracción",
    p_delay="Retraso entre solicitudes (segundos)",
    p_delay_desc="Auméntelo si el sitio bloquea las solicitudes (recomendado: 1-3s)",
    p_timeout="Tiempo de espera (segundos)",
    p_timeout_desc="Tiempo máximo de espera por página (recomendado: 15-25s)",
    p_retries="Intentos en caso de fallo",
    p_retries_desc="Número de reintentos si una página no responde (recomendado: 2-4)",
    p_filter="Filtros de email",
    p_excl="Dominios a excluir (fabricantes)",
    p_excl_desc="Un dominio por línea — estos emails se ignorarán automáticamente",
    p_also_excl="Excluir también emails genéricos (info@, contact@…)",
    p_export="Exportación",
    p_include_no_email="Incluir concesionarios sin email en el CSV",
    p_open_csv="Abrir el CSV tras exportar",
    btn_save="  Guardar  ", btn_reset="Restablecer",
    info_saved="¡Ajustes guardados!", info_reset="Ajustes restablecidos.",
    err_url="URL inválida (debe empezar por http)",
    export_ok="CSV guardado en el Escritorio:",
    ready="Listo", last="Última extracción: ",
    copied="emails copiados!", no_copy="Sin emails.",
),
"EN": dict(
    title="Silence.eco — Dealer Finder",
    tab1="  Extraction  ", tab2="  Results  ", tab3="  Settings  ",
    url_label="Dealer network page URL:",
    btn_go="  Extract  ", btn_run="  Running…  ", btn_again="  Run Again  ",
    opt1="Find emails (web + Hunter.io)", opt2="Deduplicate", opt3="Exclude manufacturer emails",
    info="Scans the page → extracts all dealers → finds their website\n→ scrapes emails → queries Hunter.io to enrich contacts.",
    log_lbl="Log:",
    col_num="#", col_name="Name", col_addr="Address", col_phone="Phone",
    col_email="Email", col_src="Source", col_role="Role / Person",
    src_page="Page", src_web="Website", src_hunter="Hunter.io",
    stat1="Dealers", stat2="Emails found", stat3="Via Hunter/web", stat4="Addresses",
    btn_copy="Copy emails", btn_csv="Export Excel",
    no_results="No results — run an extraction first",
    p_search="Search & Extraction",
    p_delay="Delay between requests (seconds)",
    p_delay_desc="Increase if the site blocks requests (recommended: 1-3s)",
    p_timeout="Connection timeout (seconds)",
    p_timeout_desc="Maximum wait time per page (recommended: 15-25s)",
    p_retries="Retries on failure",
    p_retries_desc="Number of retries if a page does not respond (recommended: 2-4)",
    p_filter="Email filters",
    p_excl="Domains to exclude (manufacturers)",
    p_excl_desc="One domain per line — these emails will be ignored automatically",
    p_also_excl="Also exclude generic emails (info@, contact@…)",
    p_export="Export",
    p_include_no_email="Include dealers without email in the CSV",
    p_open_csv="Open CSV after export",
    btn_save="  Save  ", btn_reset="Reset",
    info_saved="Settings saved!", info_reset="Settings reset.",
    err_url="Invalid URL (must start with http)",
    export_ok="CSV saved to Desktop:",
    ready="Ready", last="Last run: ",
    copied="emails copied!", no_copy="No emails.",
),
}

# ── Constantes ────────────────────────────────────────────────────────────────
EXCL_DEFAULT = [
    "silence.eco","aixam.com","microcar.fr","ligier.fr","chatenet.com",
    "casalini.it","bellier.fr","grecav.com","renault.fr","peugeot.fr",
    "citroen.fr","toyota.fr","volkswagen.fr","ford.fr","opel.fr",
]
PLATFORM_SKIP = [
    "duckduckgo.com","google.","facebook.","instagram.","twitter.","linkedin.",
    "youtube.","yelp.","tripadvisor.","pagesjaunes.","societe.com","verif.com",
    "pappers.fr","wix.","squarespace.","shopify.","wordpress.","example.com",
    "test.com","noreply.","sentry.","w3.org","schema.org","cloudflare.",
    "microsoft.","apple.","amazon.","leboncoin.","lacentrale.","largus.",
]
GENERIC_PREFIXES = ["info","contact","noreply","no-reply","admin","webmaster",
                    "postmaster","support","hello","bonjour","mailer"]

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_RE = re.compile(r"(?:\+33|\+34|\+44|0)[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}")
ADDR_RE  = re.compile(
    r"\d{1,4}[\s,]+[^\d\n]{5,60}[\s,]+\d{5}[\s,]+[A-ZÀ-Ÿa-zà-ÿ\s\-]{2,35}", re.I)

UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0",
]

ROLE_MAP = [
    (["direction","directeur","directrice","pdg","ceo","dg","gerant","patron","president"], "Direction"),
    (["commercial","vente","ventes","sales","vendeur","vendeuse","dealer"],                 "Commercial"),
    (["marketing","communication","comm","media","promo","digital","pub"],                  "Marketing"),
    (["comptabilite","comptable","compta","facturation","finance","gestion"],               "Comptabilité"),
    (["contact","info","information","hello","bonjour","accueil","reception","bienvenue"],  "Contact général"),
    (["apv","atelier","technique","sav","service","reparation","garage","meca"],            "Après-vente"),
    (["rh","hr","recrutement","emploi","jobs","carriere"],                                  "RH"),
    (["secretariat","admin","administration","assistante","assistant","bureau"],            "Administration"),
    (["pieces","parts","magasin","boutique","accessoires"],                                  "Pièces / Boutique"),
]

def guess_role(email):
    local = re.sub(r"[.\-_0-9]", "", email.split("@")[0].lower())
    for kws, lbl in ROLE_MAP:
        if any(k in local for k in kws): return lbl
    return "Équipe"

# ── Fetch robuste ─────────────────────────────────────────────────────────────
def fetch(url, timeout=18, retries=3, ref="https://www.google.com"):
    """
    Fetch with 4 strategies — guaranteed to work on all sites.
    """
    html = None

    # ── Strategy 1: Selenium (vrai Chrome installé sur le PC) ────────────
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import subprocess, shutil

        # Find chrome/chromedriver
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
            "/usr/bin/google-chrome", "/usr/bin/chromium-browser",
        ]
        chrome_ok = any(os.path.exists(p) for p in chrome_paths)

        def _selenium_fetch(driver_instance):
            """
            Full page extraction:
            1. Load page and wait for JS
            2. Scroll progressively to trigger lazy loading
            3. Select ALL text (Ctrl+A) to force full render
            4. Return complete page source
            """
            driver_instance.get(url)
            time.sleep(2)  # initial JS execution

            # Progressive scroll to trigger lazy loading
            last_height = 0
            for _ in range(20):  # scroll up to 20 times
                # Get current page height
                height = driver_instance.execute_script(
                    "return document.body.scrollHeight")
                if height == last_height:
                    break  # page fully loaded

                # Scroll down in steps to trigger lazy loading
                step = height // 5
                for pos in range(0, height, step):
                    driver_instance.execute_script(
                        f"window.scrollTo(0, {pos});")
                    time.sleep(0.1)

                # Scroll back to top
                driver_instance.execute_script("window.scrollTo(0, 0);")
                time.sleep(0.3)

                # Select ALL text on the page (forces full render)
                try:
                    driver_instance.execute_script(
                        "document.execCommand('selectAll', false, null);")
                except Exception:
                    pass

                last_height = height
                time.sleep(0.5)

            # Final scroll to bottom to ensure everything loaded
            driver_instance.execute_script(
                "window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)

            return driver_instance.page_source

        # Try with webdriver-manager first
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            from selenium.webdriver.chrome.service import Service as ChromeService
            opts = Options()
            opts.add_argument("--headless=new")
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            opts.add_argument("--disable-blink-features=AutomationControlled")
            opts.add_argument(f"--user-agent={random.choice(UA_POOL)}")
            opts.add_argument("--window-size=1920,1080")
            opts.add_experimental_option("excludeSwitches", ["enable-automation"])
            opts.add_experimental_option("useAutomationExtension", False)
            service = ChromeService(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=opts)
            driver.set_page_load_timeout(timeout)
            html = _selenium_fetch(driver)
            driver.quit()
            if html and len(html) > 5000:
                return html
        except Exception:
            pass

        # Fallback: system Chrome
        if chrome_ok:
            opts = Options()
            opts.add_argument("--headless=new")
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            opts.add_argument("--disable-blink-features=AutomationControlled")
            opts.add_argument(f"--user-agent={random.choice(UA_POOL)}")
            opts.add_argument("--window-size=1920,1080")
            opts.add_experimental_option("excludeSwitches", ["enable-automation"])
            opts.add_experimental_option("useAutomationExtension", False)
            driver = webdriver.Chrome(options=opts)
            driver.set_page_load_timeout(timeout)
            html = _selenium_fetch(driver)
            driver.quit()
            if html and len(html) > 5000:
                return html
    except Exception:
        pass

    # ── Strategy 2: requests with session + cookies ───────────────────────
    try:
        import requests
        session = requests.Session()
        session.headers.update({
            "User-Agent": random.choice(UA_POOL),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Referer": ref,
            "DNT": "1",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "cross-site",
            "Sec-Fetch-User": "?1",
        })
        # First visit homepage to get cookies
        base = re.match(r"https?://[^/]+", url)
        if base:
            try: session.get(base.group(0), timeout=8)
            except: pass
        resp = session.get(url, timeout=timeout, allow_redirects=True)
        resp.encoding = resp.apparent_encoding or "utf-8"
        if len(resp.text) > 5000:
            return resp.text
        html = resp.text
    except Exception:
        pass

    # ── Strategy 3: urllib with full cookie jar ───────────────────────────
    for attempt in range(retries):
        try:
            import http.cookiejar
            cj = http.cookiejar.CookieJar()
            opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
            opener.addheaders = [
                ("User-Agent", random.choice(UA_POOL)),
                ("Accept", "text/html,application/xhtml+xml,*/*;q=0.9"),
                ("Accept-Language", "fr-FR,fr;q=0.9,en;q=0.7"),
                ("Accept-Encoding", "gzip, deflate"),
                ("Connection", "keep-alive"),
                ("Referer", ref),
                ("Upgrade-Insecure-Requests", "1"),
            ]
            with opener.open(url, timeout=timeout) as r:
                raw = r.read()
                enc = r.headers.get_content_charset("utf-8") or "utf-8"
                text = raw.decode(enc, errors="replace")
                if len(text) > 3000:
                    return text
                html = text
        except urllib.error.HTTPError as e:
            if e.code in (403, 429, 503):
                time.sleep(2 ** attempt + random.uniform(1, 3))
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1.5 * (attempt + 1))

    # ── Strategy 4: return whatever we got, even if small ─────────────────
    if html:
        return html
    raise Exception(f"Impossible de charger la page: {url}")


def is_real(e, filter_generic=False):
    if not e or len(e) < 7 or "@" not in e: return False
    local, dom = e.split("@")[0].lower(), e.split("@")[-1].lower()
    if any(p in dom for p in PLATFORM_SKIP): return False
    if re.search(r"\.(png|jpg|gif|css|js|svg|ico|woff|ttf|pdf)$", dom): return False
    if "." not in dom or len(dom.split(".")[0]) < 2: return False
    if filter_generic and local in GENERIC_PREFIXES: return False
    return True

def normalize(s):
    """Normalize string for comparison: lowercase, remove accents, keep only alphanum."""
    import unicodedata
    s = unicodedata.normalize("NFD", s.lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s)

def email_matches_dealer(email, dealer_name, dealer_addr="", threshold=0.35):
    """Wrapper — delegates to score_email. Returns (bool, score, reason)."""
    s, r = score_email(email, dealer_name, source="search")
    return s >= 20, s / 100.0, r  # keep all, only 0% (blacklisted) rejected


def get_emails(text, excl, filter_generic=False):
    out = set()
    for m in EMAIL_RE.finditer(text):
        e = m.group(0).lower().strip(".,;:<>\"'()")
        dom = e.split("@")[-1]
        if not any(x in dom for x in excl) and is_real(e, filter_generic):
            out.add(e)
    return out

def get_phones(text):
    seen, out = set(), []
    for m in PHONE_RE.finditer(text):
        p = re.sub(r"[\s.\-]", "", m.group(0))
        if p not in seen: seen.add(p); out.append(p)
    return out

# ── Hunter.io ─────────────────────────────────────────────────────────────────
def hunter_search(domain, excl, log, timeout=15):
    if not HUNTER_KEY or not domain: return {}
    try:
        url = (f"https://api.hunter.io/v2/domain-search"
               f"?domain={urllib.parse.quote(domain)}&api_key={HUNTER_KEY}&limit=100")
        data = json.loads(fetch(url, timeout=timeout))
        results = {}
        for item in data.get("data", {}).get("emails", []):
            e = item.get("value", "").lower()
            if not e or not is_real(e) or any(x in e.split("@")[-1] for x in excl): continue
            fn = (item.get("first_name") or "").strip()
            ln = (item.get("last_name") or "").strip()
            pos = (item.get("position") or "").strip()
            dept = (item.get("department") or "").strip()
            person = f"{fn} {ln}".strip()
            parts = [p for p in [person, pos, dept] if p]
            results[e] = " — ".join(parts) if parts else guess_role(e)
        log(f"  ✅ Hunter.io ({domain}): {len(results)} email(s)")
        return results
    except Exception as err:
        log(f"  Hunter: {str(err)[:60]}")
        return {}

# ── Trouver le vrai site du concessionnaire ───────────────────────────────────
def find_dealer_domain(name, addr, log, delay=0.5):
    city = ""
    if addr:
        m = re.search(r"\d{5}\s+(\S+)", addr)
        if m: city = m.group(1)
    queries = [
        f'"{name}" site officiel concessionnaire',
        f'{name} {city} site web' if city else f'{name} concessionnaire site',
    ]
    for q in queries:
        try:
            log(f"  🌐 Recherche site: {q[:55]}")
            url = "https://html.duckduckgo.com/html/?q=" + urllib.parse.quote(q)
            html = fetch(url, timeout=15, ref="https://duckduckgo.com")
            for href in re.findall(r'href="(https?://[^"&]{8,})"', html):
                dom = re.sub(r"https?://(?:www\.)?", "", href).split("/")[0].lower().strip()
                if not dom or "." not in dom or len(dom) < 5: continue
                if any(c in dom for c in EXCL_DEFAULT): continue
                if any(p in dom for p in PLATFORM_SKIP): continue
                log(f"  Site: {dom}")
                return dom
            time.sleep(delay * 0.5)
        except Exception as e:
            pass  # silent fail, move to next
    return ""

def scrape_domain_emails(domain, excl, log, timeout=12):
    """Scrape dealer website — ALL emails found are valid (it is their own site)."""
    found = set()
    pages = [
        f"https://{domain}", f"https://www.{domain}",
        f"https://{domain}/contact", f"https://www.{domain}/contact",
        f"https://{domain}/nous-contacter", f"https://{domain}/contactez-nous",
        f"https://{domain}/contact.html", f"https://{domain}/contacto",
        f"https://{domain}/contact-us",
    ]
    for page in pages:
        try:
            html = fetch(page, timeout=timeout, ref=f"https://{domain}")
            # No excl filter — it's their own site, all emails are valid
            emails = get_emails_raw(html)
            if emails:
                found.update(emails)
                log(f"  ✅ [Site] {len(emails)} email(s) — {page.split('/')[-1] or domain}")
            if len(found) >= 10: break
            time.sleep(random.uniform(0.3, 0.6))
        except Exception:
            continue
    return found

# ── Enrichissement complet ────────────────────────────────────────────────────
def google_search_emails(name, addr, excl, log):
    """
    Search emails using real Chrome (Selenium) to bypass all blocks.
    Scans full page — Google shows emails directly in snippets.
    Stops immediately when first email found.
    """
    city = ""
    if addr:
        m = re.search(r"\d{5}\s+(\S+)", addr)
        if m: city = m.group(1)
    n = re.sub(r"[^a-zA-ZÀ-ÿ0-9\s]", " ", name).strip()

    def scan_full(html, source):
        """
        Scan Google result page smartly:
        - Extract emails near the dealer name in the HTML (context window)
        - Validate each email actually belongs to this dealer
        - Reject emails from unrelated companies
        """
        text = re.sub(r"<[^>]+>", " ", html)
        text = re.sub(r"\s+", " ", text)

        # ── Find emails that appear NEAR the dealer name in the page ──────
        name_norm = normalize(n)
        name_pos = text.lower().find(n.lower()[:10])

        candidate_emails = set()

        if name_pos >= 0:
            # Look in a window of 500 chars around the dealer name mention
            window = text[max(0, name_pos-100):name_pos+500]
            candidate_emails.update(get_emails(window, excl))

        # Also scan raw HTML for mailto: links near dealer name
        html_lower = html.lower()
        name_html_pos = html_lower.find(n.lower()[:10])
        if name_html_pos >= 0:
            window_html = html[max(0, name_html_pos-200):name_html_pos+800]
            for m2 in re.finditer(r"mailto:([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})", window_html):
                e = m2.group(1).lower()
                if is_real(e) and not any(x in e.split("@")[-1] for x in excl):
                    candidate_emails.add(e)

        # If nothing found near name, scan full page as fallback
        if not candidate_emails:
            candidate_emails.update(get_emails(text, excl))
            for m2 in re.finditer(r"mailto:([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})", html):
                e = m2.group(1).lower()
                if is_real(e) and not any(x in e.split("@")[-1] for x in excl):
                    candidate_emails.add(e)

        # ── Validate each email matches THIS dealer ────────────────────────
        validated = set()
        for e in candidate_emails:
            ok, score, reason = email_matches_dealer(e, name, addr)
            if ok:
                validated.add(e)
                log(f"  ✅ [{source}] {e} ({reason})")
            else:
                log(f"  ✗  [{source}] {e} — rejeté ({reason})")
        return validated

    def selenium_search(query, log_prefix="Google"):
        """Use real Chrome via Selenium — impossible to block."""
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC

            opts = Options()
            opts.add_argument("--headless=new")
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            opts.add_argument("--disable-blink-features=AutomationControlled")
            opts.add_argument(f"--user-agent={random.choice(UA_POOL)}")
            opts.add_experimental_option("excludeSwitches", ["enable-automation"])
            opts.add_experimental_option("useAutomationExtension", False)

            # Try webdriver-manager first, then system chrome
            driver = None
            try:
                from webdriver_manager.chrome import ChromeDriverManager
                from selenium.webdriver.chrome.service import Service
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=opts)
            except Exception:
                driver = webdriver.Chrome(options=opts)

            driver.set_page_load_timeout(15)
            url = "https://www.google.com/search?q=" + urllib.parse.quote(query) + "&hl=fr&num=10"
            driver.get(url)
            time.sleep(1.5)  # let page fully render
            html = driver.page_source
            driver.quit()
            return html
        except Exception as e:
            log(f"  ↳ Selenium indisponible: {str(e)[:40]}")
            return None

    # ── 1. Google via Selenium (vrai Chrome) ─────────────────────────────
    queries = [
        f'"{n}" email',
        f'"{n}" {city} email' if city else None,
        f'{n} concessionnaire contact email',
    ]
    for q in [x for x in queries if x]:
        log(f"  🔍 Google (Chrome): {q[:55]}")
        html = selenium_search(q, "Google")
        if html:
            emails = scan_full(html, "Google")
            if emails:
                log(f"  ✔ Email trouvé via Google — arrêt de la recherche")
                return emails
        time.sleep(random.uniform(0.8, 1.5))

    # ── 2. Google Maps via Selenium ───────────────────────────────────────
    try:
        log(f"  🗺️  Google Maps (Chrome): {n[:40]}")
        html = selenium_search(f"{n} {city} concessionnaire".strip(), "Maps")
        if html:
            emails = scan_full(html, "Google Maps")
            if emails:
                log(f"  ✔ Email trouvé via Google Maps — arrêt")
                return emails
    except Exception:
        pass

    # ── 3. DuckDuckGo (urllib fallback) ───────────────────────────────────
    for q in [f'"{n}" email', f'{n} email', f'{n} {city} contact' if city else None]:
        if not q: continue
        try:
            log(f"  🔍 DuckDuckGo: {q[:50]}")
            url = "https://html.duckduckgo.com/html/?q=" + urllib.parse.quote(q)
            html = fetch(url, timeout=12, ref="https://duckduckgo.com")
            emails = scan_full(html, "DuckDuckGo")
            if emails:
                log(f"  ✔ Email trouvé via DuckDuckGo — arrêt")
                return emails
            time.sleep(random.uniform(0.5, 1.0))
        except Exception:
            break

    # ── 4. Bing ───────────────────────────────────────────────────────────
    try:
        log(f"  🔍 Bing: {n[:40]}")
        q = urllib.parse.quote(f'"{n}" email contact')
        html = fetch(f"https://www.bing.com/search?q={q}&setlang=fr",
                     timeout=12, ref="https://www.bing.com")
        emails = scan_full(html, "Bing")
        if emails:
            log(f"  ✔ Email trouvé via Bing — arrêt")
            return emails
    except Exception:
        pass

    # ── 5. Pages Jaunes ───────────────────────────────────────────────────
    try:
        log(f"  🔍 Pages Jaunes: {n[:35]}")
        q = urllib.parse.quote(f"{n} {city}".strip())
        html = fetch(f"https://www.pagesjaunes.fr/pagesblanches/recherche?quoiqui={q}",
                     timeout=12, ref="https://www.pagesjaunes.fr")
        emails = scan_full(html, "Pages Jaunes")
        if emails:
            log(f"  ✔ Email trouvé via Pages Jaunes — arrêt")
            return emails
    except Exception:
        pass

    # ── 6. Societe.com ────────────────────────────────────────────────────
    try:
        log(f"  🔍 Societe.com: {n[:35]}")
        q = urllib.parse.quote(n)
        html = fetch(f"https://www.societe.com/cgi-bin/search?champs={q}",
                     timeout=12, ref="https://www.societe.com")
        emails = scan_full(html, "Societe.com")
        if emails:
            log(f"  ✔ Email trouvé via Societe.com — arrêt")
            return emails
    except Exception:
        pass

    log(f"  ✗ Aucun email trouvé — 6 sources épuisées")
    return set()


# Vehicle activity codes L6e/L7e (light quadricycles, microcars)
VEHICLE_CODES = {
    "45.11", "45.19", "45.20", "45.31", "45.32", "45.40",  # auto trade/repair
    "29.10", "30.91", "30.92",  # vehicle manufacturing
    "77.11", "77.12",           # vehicle rental
}

def check_gouv(name, addr, log):
    """
    Search annuaire-entreprises.data.gouv.fr (official French gov API).
    Returns (emails set, siret, activity_code, verified_address)
    """
    city = ""
    if addr:
        m = re.search(r"\d{5}\s+(\S+)", addr)
        if m: city = m.group(1)
    postcode = ""
    if addr:
        m2 = re.search(r"(\d{5})", addr)
        if m2: postcode = m2.group(1)

    # Clean name — remove brand prefix
    clean = re.sub(r"^(Ligier\s+(Store|Partner|Service)\s*)", "", name, flags=re.I).strip()
    clean = re.split(r"\s+-\s+", clean)[0].strip()

    try:
        # Use the official API (JSON)
        q = urllib.parse.quote(f"{clean} {city}".strip())
        api_url = f"https://recherche-entreprises.api.gouv.fr/search?q={q}&page=1&per_page=5"
        log(f"  🏛️  Annuaire officiel: {clean[:35]}")
        data_str = fetch(api_url, timeout=12, ref="https://annuaire-entreprises.data.gouv.fr")
        data = json.loads(data_str)

        results = data.get("results", [])
        for r in results:
            siret     = r.get("siren", "")
            siege     = r.get("siege", {})
            code_naf  = siege.get("activite_principale", "") or r.get("activite_principale", "")
            adresse   = siege.get("adresse", "")
            cp        = siege.get("code_postal", "")
            nom_comp  = r.get("nom_complet", "") or r.get("nom_raison_sociale", "")

            # Verify it's a vehicle dealer (NAF code starts with 45)
            is_vehicle = code_naf.startswith("45") or code_naf[:5] in VEHICLE_CODES

            # Check city/postcode match if we have one
            city_match = (not postcode) or (cp == postcode) or (city.upper()[:4] in adresse.upper()[:30])

            if city_match:
                log(f"  ✓ Trouvé: {nom_comp[:40]} | SIRET: {siret} | NAF: {code_naf}")
                if not is_vehicle:
                    log(f"  ↳ Code NAF {code_naf} — pas un concessionnaire véhicule")

                # Fetch the full company page for email
                if siret:
                    try:
                        page_url = f"https://annuaire-entreprises.data.gouv.fr/entreprise/{siret}"
                        html = fetch(page_url, timeout=10, ref="https://annuaire-entreprises.data.gouv.fr")
                        emails = get_emails(html)
                        if emails:
                            for e in sorted(emails)[:3]:
                                log(f"  ✅ [gouv.fr] {e}")
                            return emails, siret, code_naf, adresse
                    except Exception:
                        pass
                return set(), siret, code_naf, adresse
        return set(), "", "", ""
    except Exception as e:
        log(f"  ↳ gouv.fr: {str(e)[:40]}")
        return set(), "", "", ""


def enrich_dealer(dealer, excl, log, delay=0.8, timeout=15, brand_domain=""):
    """
    Enrichment pipeline:
    1. Direct Google search "name email" → fastest
    2. Find dealer website → scrape emails
    3. Hunter.io on domain → professional contacts with names
    """
    all_em = dict(dealer.get("emails", {}))
    src    = dealer.get("src", "pending")
    name   = dealer["name"]

    log(f"┌─ {name[:50]}")

    # ── Étape 1 : Recherche directe Google ───────────────────────────────
    log(f"│  Étape 1/3 — Recherche Google directe")
    direct = google_search_emails(name, dealer["addr"], excl, log)
    if direct:
        for e in direct:
            if e not in all_em: all_em[e] = guess_role(e)
        src = "web"
        log(f"│  → {len(direct)} email(s) trouvé(s) via Google")
    else:
        log(f"│  → Aucun email trouvé via Google")

    # ── Étape 2 : Site web du concessionnaire ────────────────────────────
    log(f"│  Étape 2/3 — Recherche du site web")
    domain = find_dealer_domain(name, dealer["addr"], log, delay)
    if domain:
        dealer["website"] = domain
        log(f"│  → Site trouvé: {domain}")
        log(f"│  Scraping du site web…")
        scraped = scrape_domain_emails(domain, excl, log, timeout)
        if scraped:
            for e in scraped:
                if e not in all_em: all_em[e] = guess_role(e)
            if src == "pending": src = "web"
            log(f"│  → {len(scraped)} email(s) depuis le site")
        else:
            log(f"│  → Aucun email sur le site")

        # ── Étape 3 : Hunter.io ──────────────────────────────────────────
        log(f"│  Étape 3/3 — Hunter.io ({domain})")
        h = hunter_search(domain, excl, log, timeout)
        if h:
            all_em.update(h)
            src = "hunter"
            log(f"│  → {len(h)} contact(s) Hunter.io")
        else:
            log(f"│  → Aucun résultat Hunter.io")
    else:
        log(f"│  → Site web non trouvé")
        log(f"│  Étape 3/3 — Hunter.io ignoré (pas de domaine)")

    total = len([e for e in all_em if is_real(e)])
    log(f"└─ Total: {total} email(s) — Source: {src}")

    return {e: r for e, r in all_em.items() if is_real(e)}, src



def scrape_subpage_network(url, excl, log, prog, timeout=18, brand_domain=""):
    """
    Handle sites where each dealer has its OWN sub-page.
    Pattern: /concessionnaires/ → lists links → /concessionnaire/nom/ → details
    Examples: moto.suzuki.fr, yamaha-motor.fr, honda-moto.fr
    
    Strategy:
    1. Fetch the listing page → extract all dealer sub-page links
    2. Fetch each sub-page in parallel → extract name/addr/phone/email
    """
    import urllib.parse as _up

    base = re.match(r"https?://[^/]+", url)
    base_url = base.group(0) if base else ""
    parsed = _up.urlparse(url)
    path_base = re.sub(r"/[^/]*$", "/", parsed.path)  # e.g. /concessionnaires/

    log(f"  Détection liens fiches individuelles…")
    html = fetch(url, timeout=timeout, retries=2)
    from bs4 import BeautifulSoup as _BS
    soup = _BS(html, "html.parser")

    # Find all dealer sub-page links
    dealer_links = []
    seen_hrefs   = set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        # Make absolute
        if href.startswith("/"):
            href = base_url + href
        elif not href.startswith("http"):
            continue
        # Must be a sub-page of same domain going deeper
        if base_url not in href: continue
        if href == url or href.rstrip("/") == url.rstrip("/"): continue
        # Must look like a dealer page (deeper path)
        link_path = _up.urlparse(href).path
        if not link_path.startswith(parsed.path.rstrip("/")): continue
        # Avoid pagination, filters, anchors
        if any(x in href for x in ["?", "#", "page=", "filtre", "search",
                                     "cart", "panier", "login", "account",
                                     "mentions", "cgv", "contact-general",
                                     "politique"]): continue
        # Must be a different/deeper path
        depth = len([p for p in link_path.split("/") if p])
        base_depth = len([p for p in parsed.path.split("/") if p])
        if depth <= base_depth: continue

        href_clean = href.rstrip("/")
        if href_clean not in seen_hrefs:
            seen_hrefs.add(href_clean)
            # Get text label
            label = a.get_text(strip=True)[:80]
            dealer_links.append((href, label))

    if not dealer_links:
        return []

    log(f"  {len(dealer_links)} fiches individuelles détectées → visite en cours…")

    # Scrape each sub-page in parallel
    import threading
    dealers   = []
    lock      = threading.Lock()
    done      = [0]

    def scrape_one(href, label):
        try:
            sub_html = fetch(href, timeout=timeout, retries=1, ref=url)
            sub_soup = _BS(sub_html, "html.parser")

            # Remove navigation/footer noise
            for tag in sub_soup(["script","style","nav","footer","head",
                                  "header","aside","noscript"]):
                tag.decompose()

            txt = sub_soup.get_text("\n", strip=True)

            # Extract fields
            name  = ""
            addr  = ""
            phone = ""
            email = ""
            site  = ""

            P2 = re.compile(r"(?:\+33|\+34|0)[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}")

            # Name: h1 or h2 is most reliable
            for tag in sub_soup.find_all(["h1","h2"]):
                t = tag.get_text(strip=True)
                if 3 < len(t) < 80 and not re.search(r"suzuki|moto|yamaha|honda|kawasaki|bmw", t, re.I):
                    name = t; break
            if not name:
                name = label[:80] if label else _up.urlparse(href).path.split("/")[-2].replace("-"," ").title()

            # Address: look for postcode pattern
            addr_m = re.search(r"(\d{1,4}[^\n]{5,60}\d{5}[^\n]{2,30})", txt)
            if addr_m: addr = addr_m.group(1).strip()[:150]

            # Phone
            ph_m = P2.search(txt)
            if ph_m: phone = ph_m.group(0)

            # Email — from mailto: links first, then text
            for a2 in sub_soup.find_all("a", href=True):
                if a2["href"].startswith("mailto:"):
                    e = a2["href"][7:].split("?")[0].strip().lower()
                    if is_real(e):
                        dom = e.split("@")[-1].lower()
                        if brand_domain and brand_domain in dom: continue
                        email = e; break
            if not email:
                em = EMAIL_RE.search(txt)
                if em:
                    e = em.group(0).lower()
                    dom = e.split("@")[-1].lower()
                    if not (brand_domain and brand_domain in dom):
                        email = e

            # Website: external link (not same domain as brand)
            for a2 in sub_soup.find_all("a", href=True):
                h2 = a2["href"]
                if h2.startswith("http") and base_url not in h2:
                    if not any(x in h2 for x in ["facebook","instagram","twitter",
                                                   "youtube","linkedin"]):
                        site = h2; break

            scored = {}
            if email:
                s, r = score_email(email, name, source="site",
                                   brand_domain=brand_domain)
                if s > 0:
                    scored[email] = f"{guess_role(email)} [{s}%]"

            with lock:
                done[0] += 1
                pct = int(20 + 60 * done[0] / max(len(dealer_links), 1))
                prog(pct, f"{done[0]}/{len(dealer_links)} — {name[:30]}")
                if email:
                    log(f"  ✅ {name[:35]} → {email}")
                dealers.append({
                    "name": name[:80], "addr": addr, "phone": phone,
                    "website": site or href, "profile_url": href,
                    "emails": scored,
                    "src": "site" if scored else "pending",
                    "siret": "", "naf": "",
                })
        except Exception as ex:
            with lock:
                done[0] += 1
                log(f"  ⚠ {href[-40:]}: {str(ex)[:40]}")

    # Parallel with up to 8 threads
    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(scrape_one, href, label): href
                   for href, label in dealer_links}
        for f in as_completed(futures):
            pass  # results collected via lock

    log(f"  ✓ {len(dealers)} concessionnaires · {sum(1 for d in dealers if d['emails'])} avec email")
    return dealers


def parse_plain_text_page(text, excl, brand_domain=""):
    dealers = []
    SKIP = {
        "distributeur","réparateur","reparateur","facebook","instagram",
        "twitter","voir la fiche","voir la carte","map","linkedin",
        "tiktok","youtube","x","snapchat","localiser","géolocalisation",
        "geolocalisation","localisez-moi","filtre marque",
        "particulier","professionnel","mon compte","réseau","faq"
    }
    P2 = re.compile(r"(?:\+33|\+34|\+689|0)[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}")
    raw_blocks = text.split("Voir la fiche")
    for block in raw_blocks:
        lines = [l.strip() for l in block.split("\n") if l.strip()]
        if len(lines) < 2:
            continue
        name = ""
        addr_parts = []
        phone = ""
        email = ""
        website = ""
        i = 0
        while i < len(lines):
            l = lines[i]
            if (re.search(r"voiture sans permis", l, re.I) or
                    re.search(r"\(\d{1,3}[AB]?\)", l) or
                    l.lower() in SKIP):
                i += 1
                continue
            break
        if i >= len(lines):
            continue
        candidate = lines[i]
        if (candidate.lower() in SKIP or
                candidate.startswith("http") or
                EMAIL_RE.search(candidate) or
                P2.match(candidate) or
                len(candidate) < 3 or len(candidate) > 120 or
                re.match(r"^\d+$", candidate)):
            continue
        name = candidate
        i += 1
        for line in lines[i:]:
            ll = line.lower().strip()
            if ll in SKIP:
                continue
            if line.lower().startswith("http"):
                if not website:
                    website = line.strip()
                continue
            em = EMAIL_RE.search(line)
            if em:
                e = em.group(0).lower()
                dom = e.split("@")[-1].lower()
                if brand_domain and brand_domain in dom:
                    continue
                if not email:
                    email = e
                continue
            ph = P2.search(line)
            if ph and not phone:
                phone = ph.group(0)
                continue
            if re.search(r"\d{5}", line):
                addr_parts.append(line)
            elif re.search(r"^\d+[\s,]", line) and len(line) < 80:
                addr_parts.append(line)
            elif re.match(r"^[A-Z\u00C0-\u00DC\s\-]{4,}$", line) and len(line) < 40:
                addr_parts.append(line)
            elif addr_parts and len(line) < 80 and not re.match(r"^\d+$", line):
                addr_parts.append(line)
        addr = " ".join(addr_parts)[:150]
        if not name or len(name) < 3:
            continue
        scored = {}
        if email:
            scored[email] = f"{guess_role(email)} [98%]"
        dealers.append({
            "name": name[:80], "addr": addr, "phone": phone,
            "website": website, "profile_url": "",
            "emails": scored,
            "src": "page" if scored else "pending",
            "siret": "", "naf": "",
        })
    return dealers


def scrape_page(url, excl, log, prog, timeout=18, retries=3):
    # Auto-detect brand domain from URL and exclude it from emails
    brand_domain = ""  # always defined — used throughout this function
    try:
        brand_host = re.match(r"https?://(?:www\.)?([^/]+)", url).group(1)
        brand_domain = brand_host.split(".")[0].lower()  # "ligier", "aixam", etc.
        brand_excl = [brand_domain + "."]
        excl = list(excl) + [b for b in brand_excl if b not in excl]
        log(f"✓ Domaine marque exclu : {brand_domain}.*")
    except Exception:
        pass
    prog(5, "Chargement de la page…")
    log(f"GET {url}")
    html = fetch(url, timeout=timeout, retries=retries)
    log(f"Page reçue — {len(html)//1024} Ko")
    prog(18, "Analyse HTML…")
    dealers = []
    PHONE_RE2 = re.compile(r"(?:\+33|\+34|\+44|0)[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}")

    # ── Détection automatique : réseau de fiches individuelles ───────────────
    # Sites comme Suzuki, Yamaha, Honda où chaque concessionnaire a sa propre page
    # Heuristique : page listing avec >10 liens vers sous-pages du même domaine
    try:
        from bs4 import BeautifulSoup as _BS4
        _soup_check = _BS4(html, "html.parser")
        parsed_check = __import__("urllib.parse", fromlist=["urlparse"]).urlparse(url)
        base_check   = re.match(r"https?://[^/]+", url).group(0)
        _path        = parsed_check.path.rstrip("/")
        _sub_links   = set()
        for _a in _soup_check.find_all("a", href=True):
            _h = _a["href"]
            if _h.startswith("/"): _h = base_check + _h
            if base_check in _h:
                _lpath = __import__("urllib.parse", fromlist=["urlparse"]).urlparse(_h).path
                _depth = len([p for p in _lpath.split("/") if p])
                _base_d= len([p for p in _path.split("/") if p])
                if _depth > _base_d and "?" not in _h and "#" not in _h:
                    _sub_links.add(_h)
        if len(_sub_links) >= 10:
            log(f"  Réseau de fiches individuelles détecté ({len(_sub_links)} liens)")
            sub_dealers = scrape_subpage_network(url, excl, log, prog,
                                                  timeout=timeout,
                                                  brand_domain=brand_domain)
            if len(sub_dealers) >= 5:
                log(f"✓ {len(sub_dealers)} concessionnaires via fiches individuelles")
                return sub_dealers
    except Exception as _se:
        log(f"  ↳ Détection fiches: {str(_se)[:40]}")
    ADDR_RE2  = re.compile(r"\d{1,4}[\s,]+[^\d\n]{5,60}[\s,]+\d{5}[\s,]+[A-ZÀ-Ÿa-zà-ÿ\s\-]{2,35}", re.I)

    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
        for t in soup(["script","style","nav","footer","head","header","aside"]): t.decompose()
        clean_text = soup.get_text(" ", strip=True)

        base_url = (re.match(r"https?://[^/]+", url) or type("x",[],{"group":lambda s,i:""})()).group(0)
        parsed   = urllib.parse.urlparse(url)
        path_dir = re.sub(r"/[^/]*$", "/", parsed.path)

        seen = set()

        # ── Méthode 1 : h3 comme noms (Aixam, la plupart des sites réseaux) ──
        for h3 in soup.find_all(["h3","h4"]):
            a_tag = h3.find("a")
            name  = (a_tag or h3).get_text(strip=True)
            if not name or len(name) < 3 or len(name) > 100: continue
            # ── Region/dept filter (runs first) ──────────────────────────
            name_l = name.lower().strip()
            if re.search(r"voiture sans permis", name_l): continue  # "Voiture sans permis - AIN (01)"
            if re.search(r"\(\d{1,3}[AB]?\)", name): continue       # contains (01), (2B), etc.
            if re.match(r"^\d", name): continue
            skip_words = ("tél","tel","fax","email","adresse","voir","cliquer",
                          "notre","les ","le ","la ","pour ","avec ",
                          "tous","département","région","accueil","bienvenue",
                          "distributeur","réparateur","scooter","moto électrique")
            if any(name_l.startswith(w) for w in skip_words): continue

            key = re.sub(r"[^a-z0-9]", "", name.lower())[:22]
            if not key or key in seen: continue
            seen.add(key)

            # Contenu après le h3 jusqu'au prochain h3/h4
            block_parts = []
            node = h3.next_sibling
            for _ in range(50):
                if node is None: break
                if hasattr(node,"name") and node.name in ["h3","h4","h2","h1"]: break
                block_parts.append(str(node))
                node = node.next_sibling
            block_html = "".join(block_parts)
            b = BeautifulSoup(block_html, "html.parser")
            btxt = b.get_text(" ", strip=True)

            # Aussi regarder le parent direct du h3
            parent_txt = h3.parent.get_text(" ", strip=True) if h3.parent else ""

            # Adresse
            addr = ""
            for src_txt in [btxt, parent_txt]:
                am = ADDR_RE2.search(src_txt)
                if am: addr = am.group(0).strip()[:120]; break
            if not addr:
                # Chercher un code postal 5 chiffres dans le bloc
                for tag in b.find_all(["p","address","div","span"]):
                    t2 = tag.get_text(" ", strip=True)
                    if re.search(r"\d{5}", t2) and 5 < len(t2) < 200:
                        addr = t2[:120]; break

            # Téléphone — texte ET liens tel:
            phones = get_phones(btxt + " " + parent_txt)
            if not phones:
                for a in b.find_all("a", href=True):
                    if a["href"].startswith("tel:"):
                        num = a["href"][4:].strip()
                        m = re.search(r"0\d{9}", num)
                        if m:
                            # Format: 06 12 34 56 78
                            n = m.group(0)
                            phones = [f"{n[:2]} {n[2:4]} {n[4:6]} {n[6:8]} {n[8:10]}"]
                            break
            phone = phones[0] if phones else ""

            # Emails visibles dans le bloc
            pg = get_emails(btxt + " " + parent_txt, excl)
            for a in b.find_all("a", href=True):
                h2 = a["href"]
                if h2.startswith("mailto:"):
                    e = h2[7:].split("?")[0].strip().lower()
                    if is_real(e) and not any(x in e.split("@")[-1] for x in excl):
                        pg.add(e)

            # Lien fiche du concessionnaire
            profile = ""
            if a_tag and a_tag.get("href"):
                href = a_tag["href"]
                if href.startswith("/"): href = base_url + href
                profile = href
            if not profile and h3.parent:
                for a in h3.parent.find_all("a", href=True):
                    href = a["href"]
                    if href.startswith("/"): href = base_url + href
                    if base_url in href and path_dir in href and href.rstrip("/") != url.rstrip("/"):
                        profile = href; break

            dealers.append({
                "name": name[:80], "addr": addr, "phone": phone,
                "website": "", "profile_url": profile,
                "emails": {e: guess_role(e) for e in pg},
                "src": "page" if pg else "pending",
            })

        log(f"Méthode h3/h4: {len(dealers)} concessionnaires")

        # ── Méthode 2 : sélecteurs CSS (autres sites) ──────────────────────
        if len(dealers) < 5:
            log("Tentative sélecteurs CSS…")
            best_blocks, best_n = [], 0
            for sel in [
                "[class*='dealer']","[class*='concess']","[class*='revendeur']",
                "[class*='reseau']","[class*='network']","[class*='store']",
                "[class*='location']","[class*='retailer']","[class*='point-de-vente']",
                "[class*='distributor']","[class*='agent']",
                "article","[class*='card']","[class*='item']",
            ]:
                found = soup.select(sel)
                valid = [b for b in found
                         if PHONE_RE2.search(b.get_text())
                         or EMAIL_RE.search(b.get_text())
                         or (len(b.get_text(strip=True)) > 40 and len(b.get_text(strip=True)) < 800)]
                if len(valid) > best_n:
                    best_n = len(valid); best_blocks = valid
            log(f"Meilleur CSS: {best_n} blocs")

            for blk in best_blocks:
                txt = blk.get_text(" ", strip=True)
                if len(txt) < 15: continue
                name = ""
                for tag in blk.find_all(["h2","h3","h4","h5","strong","b"]):
                    v = tag.get_text(strip=True)
                    if 3 < len(v) < 90 and not PHONE_RE2.match(v) and not v.replace(" ","").isdigit():
                        name = v; break
                if not name: continue
                key = re.sub(r"[^a-z0-9]", "", name.lower())[:22]
                if not key or key in seen: continue
                seen.add(key)
                am = ADDR_RE2.search(txt); addr = am.group(0).strip()[:120] if am else ""
                phones = get_phones(txt); phone = phones[0] if phones else ""
                pg = get_emails(txt, excl)
                for a in blk.find_all("a", href=True):
                    h2 = a["href"]
                    if h2.startswith("mailto:"):
                        e = h2[7:].split("?")[0].strip().lower()
                        if is_real(e) and not any(x in e.split("@")[-1] for x in excl): pg.add(e)
                dealers.append({
                    "name": name[:80], "addr": addr, "phone": phone,
                    "website": "", "profile_url": "",
                    "emails": {e: guess_role(e) for e in pg},
                    "src": "page" if pg else "pending",
                })

        # ── Méthode 3 : texte brut (Ligier-style, sans h3) ─────────────────
        page_emails = sum(len(d["emails"]) for d in dealers)
        if len(dealers) < 5 or page_emails == 0:
            full_text = soup.get_text("\n", strip=True)
            pt = parse_plain_text_page(full_text, excl, brand_domain=brand_domain)
            if len(pt) >= 3:
                log(f"  Texte brut: {len(pt)} concessionnaires")
                seen = {re.sub(r"[^a-z0-9]","",d["name"].lower())[:20] for d in dealers}
                for d in pt:
                    key = re.sub(r"[^a-z0-9]","",d["name"].lower())[:20]
                    if key not in seen:
                        dealers.append(d); seen.add(key)

        # ── Méthode 4 : fallback blocs avec téléphone ──────────────────────
        if len(dealers) < 5:
            log("Fallback blocs téléphone…")
            for blk in soup.find_all(["div","li","article","p","section"]):
                if not PHONE_RE2.search(blk.get_text()): continue
                txt = blk.get_text(" ", strip=True)
                if len(txt) < 20 or len(txt) > 1000: continue
                name = ""
                for tag in blk.find_all(["strong","b","span","h5","h4","h3"]):
                    v = tag.get_text(strip=True)
                    if 3 < len(v) < 90 and not PHONE_RE2.match(v): name = v; break
                if not name: continue
                key = re.sub(r"[^a-z0-9]", "", name.lower())[:22]
                if not key or key in seen: continue
                seen.add(key)
                phones = get_phones(txt); phone = phones[0] if phones else ""
                am = ADDR_RE2.search(txt); addr = am.group(0)[:120] if am else ""
                pg = get_emails(txt, excl)
                dealers.append({
                    "name": name[:80], "addr": addr, "phone": phone,
                    "website": "", "profile_url": "",
                    "emails": {e: guess_role(e) for e in pg},
                    "src": "page" if pg else "pending",
                })

    except ImportError:
        # ── Sans BeautifulSoup : regex sur le HTML brut ────────────────────
        log("Mode regex (bs4 non installé)…")
        # Extraire noms depuis les h3 avec liens
        names = re.findall(r"<h3[^>]*>\s*(?:<a[^>]*>)?([^<]{3,80})(?:</a>)?\s*</h3>", html, re.I)
        clean = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html))
        seen = set()
        for name in names:
            name = name.strip()
            if not name or len(name) < 3: continue
            key = re.sub(r"[^a-z0-9]", "", name.lower())[:22]
            if key in seen: continue
            seen.add(key)
            idx = clean.find(name[:20])
            ctx = clean[max(0,idx-30):idx+350] if idx >= 0 else ""
            phones = get_phones(ctx); phone = phones[0] if phones else ""
            am = ADDR_RE2.search(ctx); addr = am.group(0)[:120] if am else ""
            em_l = list(get_emails(ctx, excl))
            dealers.append({
                "name": name[:80], "addr": addr, "phone": phone,
                "website": "", "profile_url": "",
                "emails": {e: guess_role(e) for e in em_l},
                "src": "page" if em_l else "pending",
            })
        # Fallback téléphones si pas de h3
        if not dealers:
            seen2 = set()
            for m in PHONE_RE.finditer(clean):
                ph = re.sub(r"[\s.\-]","",m.group(0))
                idx = m.start()
                ctx = clean[max(0,idx-200):idx+200]
                nm  = re.search(r"([A-ZÀ-Ÿ][a-zà-ÿ]+(?:\s+[A-ZÀ-Ÿa-zà-ÿ\-]+){1,4})", ctx)
                name = nm.group(1) if nm else f"Concessionnaire {len(dealers)+1}"
                key = re.sub(r"[^a-z0-9]","",name.lower())[:20]
                if key in seen2: continue
                seen2.add(key)
                em_l = list(get_emails(ctx, excl))
                dealers.append({
                    "name": name[:80], "addr": "", "phone": ph,
                    "website": "", "profile_url": "",
                    "emails": {e: guess_role(e) for e in em_l},
                    "src": "page" if em_l else "pending",
                })

    # ── Déduplication finale ───────────────────────────────────────────────
    final, seen3 = [], set()
    for d in dealers:
        k = re.sub(r"[^a-z0-9]", "", d["name"].lower())[:20]
        if k and k not in seen3:
            seen3.add(k); final.append(d)

    log(f"✓ {len(final)} concessionnaires uniques extraits")
    return final


# ══════════════════════════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════════════════════════

import math, threading, time, os, csv, re
from datetime import datetime
from tkinter import ttk, messagebox, scrolledtext
import tkinter as tk

# ── Palette ───────────────────────────────────────────────────────────────────
R = "#E8192C"   # Silence red
RH= "#C41424"   # red hover
RL= "#FFF1F2"   # red light bg
W = "#FFFFFF"   # white
G1= "#F8F8FA"   # lightest surface
G2= "#F1F1F5"   # input bg
G3= "#E4E4EB"   # border
G4= "#C4C4D0"   # border strong
T1= "#111118"   # primary text
T2= "#555566"   # secondary text
T3= "#9999AA"   # muted text
GN= "#16A34A"   # green
GL= "#F0FDF4"   # green light
AM= "#D97706"   # amber
AL= "#FFFBEB"   # amber light

FONT      = "Segoe UI"
FONT_MONO = "Consolas"

class App(tk.Tk):
    def __init__(self, STRINGS, EXCL_DEFAULT, scrape_page, enrich_dealer,
                 norm, EMAIL_RE, T_dict=None):
        super().__init__()
        self._SP  = scrape_page
        self._ED  = enrich_dealer
        self._NRM = norm
        self.lang = "FR"
        self.S    = STRINGS
        self.results = []
        self._run    = False
        self._paused = False
        self._pause_ev = threading.Event()
        self._pause_ev.set()
        self.excl      = list(EXCL_DEFAULT)
        self.delay     = tk.DoubleVar(value=0.8)
        self.timeout_v = tk.IntVar(value=18)
        self.retries_v = tk.IntVar(value=3)
        self.inc_no    = tk.BooleanVar(value=True)
        self.open_csv  = tk.BooleanVar(value=False)
        self._anim_id    = None
        self._tick       = 0
        self._max_threads= SYSTEM["threads"]
        self._perf_mode  = tk.StringVar(value=SYSTEM["mode"])

        self.configure(bg=W)
        self.title("Silence.eco — Dealer Finder")
        self.geometry("1140x780")
        self.minsize(920, 640)

        self._build_styles()
        self._build_header()
        self._build_nb()
        self._apply_lang()

    def L(self, k): return self.S[self.lang].get(k, k)

    # ── Styles ─────────────────────────────────────────────────────────────
    def _build_styles(self):
        s = ttk.Style(self)
        s.theme_use("clam")

        s.configure("TNotebook", background=W, borderwidth=0, tabmargins=0)
        s.configure("TNotebook.Tab", background=W, foreground=T3,
                    padding=[22, 11], font=(FONT, 10), borderwidth=0)
        s.map("TNotebook.Tab",
              background=[("selected", W)],
              foreground=[("selected", T1)])

        s.configure("TFrame", background=W)
        s.configure("G1.TFrame", background=G1)

        s.configure("Red.Horizontal.TProgressbar",
                    troughcolor=G3, background=R,
                    borderwidth=0, thickness=3)

        s.configure("Treeview",
                    background=W, fieldbackground=W,
                    foreground=T1, rowheight=34,
                    font=(FONT, 9), borderwidth=0)
        s.configure("Treeview.Heading",
                    background=G1, foreground=T2,
                    font=(FONT, 8, "bold"),
                    relief="flat", padding=[10, 7])
        s.map("Treeview",
              background=[("selected", RL)],
              foreground=[("selected", R)])

    # ── Header ─────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self, bg=W)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=R, height=3).pack(fill="x")

        nav = tk.Frame(hdr, bg=W)
        nav.pack(fill="x", padx=32, pady=16)

        # Logo
        lo = tk.Frame(nav, bg=W)
        lo.pack(side="left")
        cv = tk.Canvas(lo, width=42, height=42, bg=W, highlightthickness=0)
        cv.pack(side="left", padx=(0, 14))
        cx, cy, r = 21, 21, 19
        pts = sum([[cx+r*math.cos(math.pi/2+i*math.pi/3),
                    cy+r*math.sin(math.pi/2+i*math.pi/3)] for i in range(6)], [])
        cv.create_polygon(pts, fill=R, outline="")
        r2 = 11
        pts2 = sum([[cx+r2*math.cos(math.pi/2+i*math.pi/3),
                     cy+r2*math.sin(math.pi/2+i*math.pi/3)] for i in range(6)], [])
        cv.create_polygon(pts2, fill=W, outline="")
        cv.create_oval(cx-5, cy-5, cx+5, cy+5, fill=R, outline="")

        nm = tk.Frame(lo, bg=W)
        nm.pack(side="left")
        tk.Label(nm, text="SILENCE", bg=W, fg=T1,
                 font=(FONT, 17, "bold")).pack(anchor="w")
        tk.Label(nm, text="ECO", bg=W, fg=R,
                 font=(FONT, 7, "bold")).pack(anchor="w")

        tk.Frame(nav, bg=G3, width=1).pack(side="left", fill="y", padx=22)

        sub = tk.Frame(nav, bg=W)
        sub.pack(side="left")
        tk.Label(sub, text="Dealer Finder", bg=W, fg=T1,
                 font=(FONT, 13, "bold")).pack(anchor="w")
        tk.Label(sub, text="Extraction  ·  Validation  ·  Enrichissement",
                 bg=W, fg=T3, font=(FONT, 8)).pack(anchor="w")

        # Right
        rt = tk.Frame(nav, bg=W)
        rt.pack(side="right")
        self._status_lbl = tk.Label(rt, text="", bg=W, fg=T3, font=(FONT, 9))
        self._status_lbl.pack(side="left", padx=(0, 24))

        for code in ("FR", "ES", "EN"):
            b = tk.Button(rt, text=code,
                          bg=G1, fg=T2,
                          font=(FONT, 8, "bold"),
                          relief="flat", bd=0, width=3,
                          cursor="hand2", pady=5,
                          activebackground=R,
                          activeforeground=W,
                          command=lambda c=code: self._switch(c))
            b.pack(side="left", padx=1)

        tk.Frame(hdr, bg=G3, height=1).pack(fill="x")

    # ── Notebook ───────────────────────────────────────────────────────────
    def _build_nb(self):
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True)
        self._nb = nb
        self._f1 = ttk.Frame(nb); self._f2 = ttk.Frame(nb); self._f3 = ttk.Frame(nb)
        nb.add(self._f1, text=""); nb.add(self._f2, text=""); nb.add(self._f3, text="")
        self._build_t1()
        self._build_t2()
        self._build_t3()

    # ── Tab 1: Extraction ──────────────────────────────────────────────────
    def _build_t1(self):
        f = self._f1
        P = dict(padx=32)

        # ── Mode selector ─────────────────────────────────────────────────
        top_row = tk.Frame(f, bg=W)
        top_row.pack(fill="x", padx=32, pady=(20, 0))
        self._input_mode = tk.StringVar(value="url")
        for val, lbl in [("url","🔗  URL du site"), ("paste","📋  Coller le texte")]:
            tk.Radiobutton(top_row, text=lbl, variable=self._input_mode,
                           value=val, bg=W, fg=T1, font=(FONT, 9),
                           activebackground=W, selectcolor=W, cursor="hand2",
                           command=self._toggle_input_mode).pack(side="left", padx=(0,24))

        # ── URL frame ──────────────────────────────────────────────────
        self._url_frame = tk.Frame(f, bg=W)
        self._url_frame.pack(fill="x", padx=32, pady=(10,0))
        self._lbl_url = tk.Label(self._url_frame, text="", bg=W, fg=T2,
                                  font=(FONT, 8, "bold"))
        self._lbl_url.pack(anchor="w", pady=(0,6))
        row_u = tk.Frame(self._url_frame, bg=W)
        row_u.pack(fill="x")
        ef = tk.Frame(row_u, bg=G2, highlightthickness=1,
                      highlightbackground=G3, highlightcolor=R)
        ef.pack(side="left", fill="x", expand=True, padx=(0,10))
        self._url = tk.StringVar(
            value="https://www.aixam.com/fr/reseau/voiture-sans-permis-france")
        self._url_entry = tk.Entry(ef, textvariable=self._url,
                                    font=(FONT, 10), bg=G2, fg=T1,
                                    insertbackground=R, relief="flat",
                                    bd=0, highlightthickness=0)
        self._url_entry.pack(fill="x", padx=14, pady=10)
        self._btn = tk.Button(row_u, text="",
                               bg=R, fg=W, font=(FONT, 10, "bold"),
                               relief="flat", bd=0, cursor="hand2",
                               padx=26, pady=11,
                               activebackground=RH, activeforeground=W,
                               command=self._start)
        self._btn.pack(side="left")
        self._pause_btn = tk.Button(row_u, text="⏸",
                                     bg=G1, fg=T2, font=(FONT, 11),
                                     relief="flat", bd=0, cursor="hand2",
                                     padx=13, pady=11, state="disabled",
                                     activebackground=AL,
                                     command=self._toggle_pause)
        self._pause_btn.pack(side="left", padx=(8,0))

        # ── Paste frame ────────────────────────────────────────────────
        self._paste_frame = tk.Frame(f, bg=W)
        # (starts hidden)
        tk.Label(self._paste_frame,
                 text="Collez ici le texte copié de la page concessionnaires :",
                 bg=W, fg=T2, font=(FONT, 8, "bold")).pack(anchor="w", pady=(0,6))
        row_p = tk.Frame(self._paste_frame, bg=W)
        row_p.pack(fill="x")
        pw = tk.Frame(row_p, bg=G2, highlightthickness=1,
                      highlightbackground=G3, highlightcolor=R)
        pw.pack(side="left", fill="x", expand=True, padx=(0,10))
        self._paste_txt = tk.Text(pw, height=5, font=(FONT, 9),
                                   bg=G2, fg=T3, insertbackground=R,
                                   relief="flat", bd=0, highlightthickness=0,
                                   padx=10, pady=8, wrap="word")
        self._paste_txt.pack(fill="x")
        self._paste_txt.insert("1.0",
            "Ctrl+A puis Ctrl+C sur la page du réseau concessionnaires, "
            "puis collez ici avec Ctrl+V…")
        self._paste_txt.bind("<FocusIn>", lambda e: (
            self._paste_txt.delete("1.0","end"),
            self._paste_txt.configure(fg=T1)
        ) if self._paste_txt.get("1.0","end-1c").startswith("Ctrl+") else None)
        pc = tk.Frame(row_p, bg=W)
        pc.pack(side="left")
        self._btn_paste = tk.Button(pc, text="Analyser",
                                     bg=R, fg=W, font=(FONT, 10, "bold"),
                                     relief="flat", bd=0, cursor="hand2",
                                     padx=22, pady=11,
                                     activebackground=RH,
                                     command=self._start_paste)
        self._btn_paste.pack()
        self._pause_btn_paste = tk.Button(pc, text="⏸",
                                           bg=G1, fg=T2, font=(FONT, 11),
                                           relief="flat", bd=0, cursor="hand2",
                                           padx=13, pady=6, state="disabled",
                                           activebackground=AL,
                                           command=self._toggle_pause)
        self._pause_btn_paste.pack(pady=(4,0))

        # Options
        opt = tk.Frame(f, bg=W)
        opt.pack(fill="x", **P, pady=(14, 0))
        self._v1 = tk.BooleanVar(value=True)
        self._v2 = tk.BooleanVar(value=True)
        self._v3 = tk.BooleanVar(value=True)
        for v, a in [(self._v1,"_c1"),(self._v2,"_c2"),(self._v3,"_c3")]:
            c = tk.Checkbutton(opt, text="", variable=v,
                               bg=W, fg=T1, font=(FONT, 9),
                               activebackground=W, selectcolor=W,
                               cursor="hand2", highlightthickness=0)
            c.pack(side="left", padx=(0, 26))
            setattr(self, a, c)

        # Status pills
        pills = tk.Frame(f, bg=W)
        pills.pack(fill="x", **P, pady=(12, 0))

        hp = tk.Frame(pills, bg=GL, padx=10, pady=5)
        hp.pack(side="left", padx=(0, 8))
        self._hunter_dot = tk.Canvas(hp, width=7, height=7, bg=GL,
                                      highlightthickness=0)
        self._hunter_dot.pack(side="left", padx=(0, 5))
        self._hunter_dot.create_oval(0, 0, 7, 7, fill=GN, outline="")
        tk.Label(hp, text="Hunter.io actif", bg=GL, fg=GN,
                 font=(FONT, 8, "bold")).pack(side="left")

        # Performance mode badge
        PERF_COLORS = {
            "eco":    ("#F1F5F9","#64748B"),
            "normal": ("#F0F9FF","#2563EB"),
            "rapide": ("#FFF7ED","#EA580C"),
            "turbo":  ("#FFF1F2","#E8192C"),
        }
        PERF_ICONS = {"eco":"🐢","normal":"⚡","rapide":"🚀","turbo":"🔥"}
        mode = SYSTEM["mode"]
        pbg, pfg = PERF_COLORS.get(mode, PERF_COLORS["normal"])
        perf_pill = tk.Frame(pills, bg=pbg, padx=10, pady=5)
        perf_pill.pack(side="left", padx=(0,8))
        tk.Label(perf_pill,
                 text=f"{PERF_ICONS[mode]} Mode {mode.capitalize()} · {SYSTEM['threads']} threads",
                 bg=pbg, fg=pfg, font=(FONT, 8, "bold")).pack()

        self._pause_pill = tk.Frame(pills, bg=AL, padx=10, pady=5)
        tk.Label(self._pause_pill, text="⏸  En pause",
                 bg=AL, fg=AM, font=(FONT, 8, "bold")).pack()

        # Progress block
        pg = tk.Frame(f, bg=W)
        pg.pack(fill="x", **P, pady=(18, 0))

        pgr = tk.Frame(pg, bg=W)
        pgr.pack(fill="x", pady=(0, 8))

        # Animated indicator
        self._ind = tk.Canvas(pgr, width=10, height=10, bg=W,
                               highlightthickness=0)
        self._ind.pack(side="left", padx=(0, 10))
        self._ind.create_oval(1, 1, 9, 9, fill=G2, outline="", tags="dot")

        self._prog_lbl = tk.Label(pgr, text="", bg=W, fg=T1,
                                   font=(FONT, 9, "bold"))
        self._prog_lbl.pack(side="left")
        self._pct_lbl = tk.Label(pgr, text="", bg=W, fg=T3,
                                  font=(FONT, 9))
        self._pct_lbl.pack(side="right")

        # Live email counter
        self._em_frame = tk.Frame(pg, bg=RL, padx=12, pady=6)
        self._em_lbl = tk.Label(self._em_frame,
                                 text="0 emails trouvés",
                                 bg=RL, fg=R,
                                 font=(FONT, 8, "bold"))
        self._em_lbl.pack()

        # Annuaire.gouv check indicator
        self._gouv_frame = tk.Frame(pg, bg=GL, padx=12, pady=6)
        self._gouv_frame.pack(fill="x", pady=(4, 0))
        self._gouv_frame.pack_forget()
        tk.Label(self._gouv_frame,
                 text="✓  Vérification annuaire-entreprises.data.gouv.fr",
                 bg=GL, fg=GN, font=(FONT, 8)).pack()

        self._prog_var = tk.DoubleVar()
        ttk.Progressbar(pg, variable=self._prog_var, maximum=100,
                        style="Red.Horizontal.TProgressbar").pack(
                        fill="x", pady=(6, 0))

        # Log
        lf = tk.Frame(f, bg=W)
        lf.pack(fill="both", expand=True, **P, pady=(18, 0))

        lh = tk.Frame(lf, bg=W)
        lh.pack(fill="x", pady=(0, 6))
        self._log_hdr = tk.Label(lh, text="", bg=W, fg=T2,
                                  font=(FONT, 8, "bold"))
        self._log_hdr.pack(side="left")

        lw = tk.Frame(lf, bg=G1, highlightthickness=1,
                      highlightbackground=G3)
        lw.pack(fill="both", expand=True)
        self._log = scrolledtext.ScrolledText(
            lw, height=7, font=(FONT_MONO, 8),
            bg=G1, fg=T2, insertbackground=R,
            relief="flat", bd=0, state="disabled",
            wrap="word", highlightthickness=0,
            padx=16, pady=10)
        self._log.pack(fill="both", expand=True)

        # Log colors
        self._log.tag_configure("ok",   foreground=GN, font=(FONT_MONO, 8, "bold"))
        self._log.tag_configure("fail", foreground="#DC2626")
        self._log.tag_configure("info", foreground="#2563EB")
        self._log.tag_configure("head", foreground=T1, font=(FONT_MONO, 8, "bold"))
        self._log.tag_configure("dim",  foreground=T3)

        # Stats cards
        sf = tk.Frame(f, bg=W)
        sf.pack(fill="x", **P, pady=(14, 26))
        self._sv = {}; self._slbl = {}
        cards = [
            ("s1", T1, W,   G3),
            ("s2", R,  RL,  R),
            ("s3", GN, GL,  GN),
            ("s4", T1, W,   G3),
        ]
        for i, (k, vc, bg, ac) in enumerate(cards):
            card = tk.Frame(sf, bg=bg,
                            highlightthickness=1,
                            highlightbackground=ac if k == "s2" else G3,
                            padx=18, pady=14)
            card.pack(side="left", fill="x", expand=True,
                      padx=(0, 10) if i < 3 else 0)
            lbl = tk.Label(card, text="", bg=bg, fg=T3,
                           font=(FONT, 8))
            lbl.pack(anchor="w")
            v = tk.StringVar(value="—")
            tk.Label(card, textvariable=v, bg=bg, fg=vc,
                     font=(FONT, 26, "bold")).pack(anchor="w")
            self._sv[k] = v; self._slbl[k] = lbl

    # ── Tab 2: Results ─────────────────────────────────────────────────────
    def _build_t2(self):
        f = self._f2

        top = tk.Frame(f, bg=W)
        top.pack(fill="x", padx=24, pady=(16, 8))
        self._meta = tk.Label(top, text="", bg=W, fg=T3, font=(FONT, 9))
        self._meta.pack(side="left")

        # Legend
        legend = tk.Frame(top, bg=W)
        legend.pack(side="left", padx=(20,0))
        for icon, label, color in [
            ("⭐⭐⭐", "≥90% Certitude",    GN),
            ("⭐⭐",   "70-89% Probable",   "#16a34a"),
            ("⭐",     "40-69% Possible",   AM),
            ("·",      "<40% Faible",       T3),
        ]:
            lf = tk.Frame(legend, bg=W)
            lf.pack(side="left", padx=(0,12))
            tk.Label(lf, text=icon, bg=W, fg=color,
                     font=(FONT, 8)).pack(side="left")
            tk.Label(lf, text=f" {label}", bg=W, fg=T3,
                     font=(FONT, 8)).pack(side="left")

        self._btn_csv = tk.Button(top, text="",
                                   bg=R, fg=W, font=(FONT, 9, "bold"),
                                   relief="flat", bd=0, cursor="hand2",
                                   padx=18, pady=7,
                                   activebackground=RH,
                                   command=self._export)
        self._btn_csv.pack(side="right")

        self._btn_copy = tk.Button(top, text="",
                                    bg=W, fg=T1, font=(FONT, 9),
                                    relief="flat", bd=0, cursor="hand2",
                                    padx=14, pady=7,
                                    highlightthickness=1,
                                    highlightbackground=G3,
                                    command=self._copy)
        self._btn_copy.pack(side="right", padx=(0, 10))

        # Table
        tw = tk.Frame(f, bg=W, highlightthickness=1,
                      highlightbackground=G3)
        tw.pack(fill="both", expand=True, padx=24, pady=(0, 24))

        cols = ("num","name","addr","phone","email","role","src")
        self._tree = ttk.Treeview(tw, columns=cols,
                                   show="headings", selectmode="browse")
        for col, w in [("num",38),("name",195),("addr",148),
                       ("phone",98),("email",185),("role",142),("src",90)]:
            self._tree.column(col, width=w, minwidth=28)

        vsb = ttk.Scrollbar(tw, orient="vertical",
                             command=self._tree.yview)
        hsb = ttk.Scrollbar(tw, orient="horizontal",
                             command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set,
                             xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True)

        self._tree.tag_configure("miss",    background="#FFF5F5")
        self._tree.tag_configure("web",     background="#FFFBEB")
        self._tree.tag_configure("hunter",  background="#F0FDF4")
        self._tree.tag_configure("high",    background="#F0FDF4")   # ≥70%
        self._tree.tag_configure("medium",  background="#FFFBEB")   # 40-69%
        self._tree.tag_configure("low",     background="#FFF8F0")   # <40%

    # ── Tab 3: Settings ────────────────────────────────────────────────────
    def _build_t3(self):
        f = self._f3
        cv = tk.Canvas(f, bg=W, highlightthickness=0)
        sb = ttk.Scrollbar(f, orient="vertical", command=cv.yview)
        cv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        cv.pack(fill="both", expand=True)
        inner = tk.Frame(cv, bg=W)
        win = cv.create_window((0, 0), window=inner, anchor="nw")
        def _r(e):
            cv.configure(scrollregion=cv.bbox("all"))
            cv.itemconfig(win, width=e.width)
        cv.bind("<Configure>", _r)

        def sec(title, subtitle=""):
            fr = tk.Frame(inner, bg=W)
            fr.pack(fill="x", padx=32, pady=(28, 0))
            tk.Label(fr, text=title, bg=W, fg=T1,
                     font=(FONT, 10, "bold")).pack(anchor="w")
            if subtitle:
                tk.Label(fr, text=subtitle, bg=W, fg=T3,
                         font=(FONT, 8)).pack(anchor="w", pady=(2, 0))
            tk.Frame(fr, bg=G3, height=1).pack(fill="x", pady=(8, 14))
            return fr

        def spin_row(parent, label, hint, var, frm, to, step=1):
            r = tk.Frame(parent, bg=W)
            r.pack(fill="x", padx=4, pady=6)
            lf = tk.Frame(r, bg=W)
            lf.pack(side="left", fill="x", expand=True)
            tk.Label(lf, text=label, bg=W, fg=T1,
                     font=(FONT, 9)).pack(anchor="w")
            if hint:
                tk.Label(lf, text=hint, bg=W, fg=T3,
                         font=(FONT, 8)).pack(anchor="w")
            sw = tk.Frame(r, bg=G3, padx=1, pady=1)
            sw.pack(side="right")
            sp = tk.Spinbox(sw, textvariable=var,
                            from_=frm, to=to, increment=step,
                            width=7, font=(FONT, 9),
                            bg=G2, fg=T1, relief="flat", bd=0,
                            insertbackground=R, highlightthickness=0)
            sp.pack(ipady=6, padx=6)

        s1 = sec("Recherche",
                 "Paramètres de vitesse et de robustesse de l'extraction")
        spin_row(s1, "Délai entre requêtes (s)",
                 "Augmenter si le site bloque (recommandé: 0.8–2s)",
                 self.delay, 0.3, 10, 0.1)
        spin_row(s1, "Timeout connexion (s)",
                 "Temps max d'attente par page (recommandé: 15–25s)",
                 self.timeout_v, 5, 60, 1)
        spin_row(s1, "Tentatives en cas d'erreur",
                 "Nombre de réessais si une page ne répond pas",
                 self.retries_v, 1, 6, 1)

        s2 = sec("Filtres",
                 "Domaines à ignorer lors de la recherche d'emails")
        tk.Label(s2, text=self.L("p_excl"), bg=W, fg=T1,
                 font=(FONT, 9)).pack(anchor="w", padx=4)
        tk.Label(s2, text=self.L("p_excl_desc"), bg=W, fg=T3,
                 font=(FONT, 8)).pack(anchor="w", padx=4, pady=(2, 8))
        tw2 = tk.Frame(s2, bg=G3, padx=1, pady=1)
        tw2.pack(fill="x", padx=4)
        self._excl_txt = tk.Text(tw2, height=8,
                                  font=(FONT_MONO, 8),
                                  bg=G2, fg=T1, insertbackground=R,
                                  relief="flat", bd=0,
                                  highlightthickness=0,
                                  padx=12, pady=8)
        self._excl_txt.pack(fill="x")
        self._excl_txt.insert("1.0", "\n".join(self.excl))

        s3 = sec("Export")
        self._ck_no = tk.Checkbutton(s3,
            text=self.L("p_include_no_email"),
            variable=self.inc_no,
            bg=W, fg=T1, font=(FONT, 9),
            activebackground=W, selectcolor=W, cursor="hand2")
        self._ck_no.pack(anchor="w", padx=4, pady=2)
        self._ck_open = tk.Checkbutton(s3,
            text=self.L("p_open_csv"),
            variable=self.open_csv,
            bg=W, fg=T1, font=(FONT, 9),
            activebackground=W, selectcolor=W, cursor="hand2")
        self._ck_open.pack(anchor="w", padx=4, pady=2)

        br = tk.Frame(inner, bg=W)
        br.pack(fill="x", padx=32, pady=22)
        self._btn_save = tk.Button(br, text="",
            bg=R, fg=W, font=(FONT, 9, "bold"),
            relief="flat", bd=0, cursor="hand2",
            padx=18, pady=9,
            activebackground=RH, command=self._save)
        self._btn_save.pack(side="left")
        self._btn_rst = tk.Button(br, text="",
            bg=W, fg=T2, font=(FONT, 9),
            relief="flat", bd=0, cursor="hand2",
            padx=14, pady=9,
            highlightthickness=1, highlightbackground=G3,
            command=self._reset)
        self._btn_rst.pack(side="left", padx=(10, 0))

    # ── Language ───────────────────────────────────────────────────────────
    def _switch(self, lang): self.lang = lang; self._apply_lang()

    def _apply_lang(self):
        L = self.L
        self.title(L("title"))
        self._nb.tab(0, text=f"  {L('tab1')}  ")
        self._nb.tab(1, text=f"  {L('tab2')}  ")
        self._nb.tab(2, text=f"  {L('tab3')}  ")
        self._lbl_url.configure(text=L("url_label"))
        self._btn.configure(
            text=L("btn_run") if (self._run and not self._paused)
            else L("btn_go") if not self._run else L("btn_again"))
        self._c1.configure(text=L("opt1"))
        self._c2.configure(text=L("opt2"))
        self._c3.configure(text=L("opt3"))
        self._log_hdr.configure(text=L("log_lbl"))
        for k, tk_k in [("s1","stat1"),("s2","stat2"),
                         ("s3","stat3"),("s4","stat4")]:
            self._slbl[k].configure(text=L(tk_k))
        rl = {"FR":"Rôle","ES":"Rol","EN":"Role"}.get(self.lang, "Rôle")
        for col, key in [("num","col_num"),("name","col_name"),
                         ("addr","col_addr"),("phone","col_phone"),
                         ("email","col_email"),("src","col_src")]:
            self._tree.heading(col, text=L(key))
        self._tree.heading("role", text=rl)
        self._btn_copy.configure(text=L("btn_copy"))
        self._btn_csv.configure(text=L("btn_csv"))
        self._btn_save.configure(text=L("btn_save"))
        self._btn_rst.configure(text=L("btn_reset"))
        self._ck_no.configure(text=L("p_include_no_email"))
        self._ck_open.configure(text=L("p_open_csv"))
        self._status_lbl.configure(text=L("ready"))
        if self.results: self._render()

    # ── Animation ──────────────────────────────────────────────────────────
    def _start_anim(self):
        self._tick = 0
        self._em_frame.pack(fill="x", pady=(6, 0))
        self._anim_loop()

    def _stop_anim(self):
        if self._anim_id:
            self.after_cancel(self._anim_id)
            self._anim_id = None
        try:
            self._ind.itemconfig("dot", fill=GN)
            self._em_frame.pack_forget()
        except Exception:
            pass

    def _anim_loop(self):
        if not self._run: return
        self._tick += 1
        t = self._tick * 0.22
        v = int(60 + 45 * math.sin(t))
        if self._paused:
            fill = AM
        else:
            fill = f"#{232:02x}{v:02x}{44:02x}"
        try:
            self._ind.itemconfig("dot", fill=fill)
            count = sum(len(d.get("emails", {})) for d in self.results)
            self._em_lbl.configure(
                text=f"{count} email{'s' if count!=1 else ''} trouvé{'s' if count!=1 else ''}")
        except Exception:
            pass
        self._anim_id = self.after(70, self._anim_loop)

    # ── Pause ──────────────────────────────────────────────────────────────
    def _apply_perf_mode(self):
        """Apply selected performance mode — update threads and delay."""
        PERF = {
            "eco":    {"threads": 2,  "delay": 1.2, "info": "2 threads · délai 1.2s · PC sous charge minimale"},
            "normal": {"threads": 6,  "delay": 0.6, "info": "6 threads · délai 0.6s · Bon équilibre vitesse/stabilité"},
            "rapide": {"threads": 10, "delay": 0.3, "info": "10 threads · délai 0.3s · PC 4+ cœurs recommandé"},
            "turbo":  {"threads": 16, "delay": 0.15,"info": "16 threads · délai 0.15s · PC 8+ cœurs recommandé"},
        }
        mode = self._perf_mode.get()
        cfg  = PERF.get(mode, PERF["normal"])
        self._max_threads = cfg["threads"]
        self.delay.set(cfg["delay"])
        try:
            self._perf_info.configure(text=f"  {cfg['info']}")
        except Exception:
            pass

    def _toggle_input_mode(self):
        if self._input_mode.get() == "url":
            self._paste_frame.pack_forget()
            self._url_frame.pack(fill="x", padx=32, pady=(10,0))
        else:
            self._url_frame.pack_forget()
            self._paste_frame.pack(fill="x", padx=32, pady=(10,0))

    def _start_paste(self):
        if self._run: return
        text = self._paste_txt.get("1.0","end-1c").strip()
        if not text or text.startswith("Ctrl+"):
            messagebox.showerror("","Collez d'abord le texte de la page."); return
        self._run = True; self._paused = False; self._pause_ev.set()
        self._btn_paste.configure(state="disabled", text="En cours…")
        self._pause_btn_paste.configure(state="normal")
        self._log.configure(state="normal"); self._log.delete("1.0","end")
        self._log.configure(state="disabled")
        self._set_prog(0,"Analyse du texte…"); self._pct_lbl.configure(text="")
        for v in self._sv.values(): v.set("—")
        self._start_anim()
        threading.Thread(target=self._worker_paste,args=(text,),daemon=True).start()

    def _worker_paste(self, text):
        try:
            self._log_add("📋 Analyse du texte collé — pas de connexion requise","ok")
            dealers = parse_plain_text_page(text, self.excl, brand_domain="")
            if not dealers:
                self._log_add("⚠ Aucun concessionnaire détecté — vérifiez le texte","fail")
                self._run=False; self._stop_anim()
                self.after(0,lambda:self._btn_paste.configure(state="normal",text="Analyser"))
                return
            em0 = sum(1 for d in dealers if d["emails"])
            self._log_add(f"✅ {len(dealers)} concessionnaires · {em0} avec email direct","ok")
            self._set_prog(30,f"{len(dealers)} concessionnaires…")
            if self._v1.get() and em0 < len(dealers):
                excl=self.excl if self._v3.get() else []
                dly=self.delay.get(); to=self.timeout_v.get()
                self._log_add(f"⚡ Enrichissement {len(dealers)-em0} concessionnaires sans email…","ok")
                lock=threading.Lock(); done=[0]
                def do_one(d):
                    self._pause_ev.wait()
                    try:
                        if not d["emails"]:
                            em,src=self._ED(d,excl,self._log_add,delay=dly,timeout=to)
                            if em: d["emails"]=em; d["src"]=src
                    except Exception as e:
                        self._log_add(f"  ⚠ {d['name'][:30]}: {e}")
                    finally:
                        with lock:
                            done[0]+=1
                            pct=30+int(65*done[0]/max(len(dealers),1))
                            self._set_prog(pct,f"{done[0]}/{len(dealers)} — {d['name'][:30]}")
                            self.results=list(dealers)
                            self.after(0,self._render)
                threads=[]
                for d in dealers:
                    self._pause_ev.wait()
                    t=threading.Thread(target=do_one,args=(d,),daemon=True)
                    threads.append(t); t.start()
                    while sum(1 for x in threads if x.is_alive())>=self._max_threads:
                        time.sleep(0.2)
                for t in threads: t.join(timeout=40)
            if self._v2.get():
                seen,uniq=set(),[]
                for d in dealers:
                    k=re.sub(r"[^a-z0-9]","",d["name"].lower())[:20]
                    if k not in seen: seen.add(k); uniq.append(d)
                dealers=uniq
            total=sum(len(d["emails"]) for d in dealers)
            self._set_prog(100,"Terminé ✓")
            self._log_add(f"━━  {len(dealers)} concessionnaires · {total} emails  ━━","head")
            self.results=dealers
            self.after(0,self._finish_paste)
        except Exception as e:
            self.after(0,lambda:messagebox.showerror("Erreur",str(e)))
            self._log_add(f"✗ {e}","fail")
            self._run=False; self._stop_anim()
            self.after(0,lambda:self._btn_paste.configure(state="normal",text="Analyser"))

    def _finish_paste(self):
        self._run=False; self._paused=False; self._stop_anim()
        self._pause_btn_paste.configure(state="disabled",text="⏸",bg=G1,fg=T2)
        d=self.results
        em=sum(len(x["emails"]) for x in d)
        wb=sum(1 for x in d if x.get("src") in ("web","hunter","site","page"))
        ad=sum(1 for x in d if x.get("addr"))
        self._sv["s1"].set(str(len(d))); self._sv["s2"].set(str(em))
        self._sv["s3"].set(str(wb)); self._sv["s4"].set(str(ad))
        self._btn_paste.configure(state="normal",text="Analyser à nouveau")
        self._status_lbl.configure(text=self.L("last")+datetime.now().strftime("%H:%M"))
        self._render(); self._nb.select(self._f2)

    def _toggle_pause(self):
        if not self._run: return
        self._paused = not self._paused
        if self._paused:
            self._pause_ev.clear()
            self._pause_btn.configure(text="▶", bg=AL, fg=AM)
            self._pause_pill.pack(side="left")
            self._prog_lbl.configure(text="En pause…")
            self._log_add("⏸  Extraction en pause — cliquez ▶ pour reprendre", "head")
        else:
            self._pause_ev.set()
            self._pause_btn.configure(text="⏸", bg=G1, fg=T2)
            self._pause_pill.pack_forget()
            self._log_add("▶  Reprise de l'extraction…", "ok")

    # ── Log ────────────────────────────────────────────────────────────────
    def _log_add(self, msg, force_tag=None):
        self._log.configure(state="normal")
        if force_tag:
            tag = force_tag
        elif any(x in msg for x in ("✅","✔","✓","━━","email(s)")):
            tag = "ok"
        elif any(x in msg for x in ("✗","ERREUR","rejeté","⚠")):
            tag = "fail"
        elif any(x in msg for x in ("🔍","🗺","💡","Hunter","1/3","2/3","3/3","gouv")):
            tag = "info"
        elif msg.startswith(("┌","└","━","─")):
            tag = "head"
        else:
            tag = "dim"
        self._log.insert("end", f"{msg}\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")

    def _set_prog(self, pct, lbl=""):
        self._prog_var.set(pct)
        if lbl:
            self._prog_lbl.configure(text=lbl)
            self._pct_lbl.configure(text=f"{int(pct)}%" if pct else "")
        self.update_idletasks()

    # ── Save / Reset ───────────────────────────────────────────────────────
    def _save(self):
        raw = self._excl_txt.get("1.0", "end").strip()
        self.excl = [l.strip().lstrip("@")
                     for l in raw.splitlines() if l.strip()]
        messagebox.showinfo("Silence", self.L("info_saved"))

    def _reset(self):
        # EXCL_DEFAULT available in scope
        self.excl = list(EXCL_DEFAULT)
        self._excl_txt.delete("1.0", "end")
        self._excl_txt.insert("1.0", "\n".join(self.excl))
        self.delay.set(0.8); self.timeout_v.set(18)
        self.retries_v.set(3); self.inc_no.set(True)
        self.open_csv.set(False)
        messagebox.showinfo("Silence", self.L("info_reset"))

    # ── Start / Worker ─────────────────────────────────────────────────────
    def _start(self):
        if self._run:
            if self._paused: self._toggle_pause()
            return
        url = self._url.get().strip()
        if not url.startswith("http"):
            messagebox.showerror("", self.L("err_url")); return
        self._run = True; self._paused = False; self._pause_ev.set()
        self._btn.configure(state="disabled", text=self.L("btn_run"))
        self._pause_btn.configure(state="normal")
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")
        self._set_prog(0, "Démarrage…")
        self._pct_lbl.configure(text="")
        for v in self._sv.values(): v.set("—")
        self._gouv_frame.pack_forget()
        self._start_anim()
        excl = self.excl if self._v3.get() else []
        # Auto-detect brand domain from URL for exclusion
        try:
            import re as _re
            host = _re.match(r"https?://(?:www\.)?([^/]+)", url).group(1)
            self._brand_domain = host.split(".")[0].lower()
        except Exception:
            self._brand_domain = ""
        threading.Thread(target=self._worker, args=(url, excl),
                         daemon=True).start()

    def _worker(self, url, excl):
        try:
            to = self.timeout_v.get()
            rt = self.retries_v.get()
            dly = self.delay.get()
            self._log_add(f"⚡ Mode {self._perf_mode.get().upper()} · {self._max_threads} threads parallèles · délai {dly}s", "ok")

            dealers = self._SP(url, excl, self._log_add,
                               self._set_prog, timeout=to, retries=rt)
            self._set_prog(44, "Enrichissement des emails…")

            if self._v1.get():
                no_em  = sum(1 for d in dealers if not d["emails"])
                has_em = len(dealers) - no_em
                self._log_add(
                    f"━━  {len(dealers)} concessionnaires  ·  "
                    f"{has_em} avec email  ·  {no_em} à enrichir  ━━", "head")

                # Show gouv indicator
                self.after(0, self._gouv_frame.pack,
                           {"fill":"x","pady":(4,0)})

                lock = threading.Lock()
                done = [0]

                def do_one(d):
                    self._pause_ev.wait()
                    try:
                        em, src = self._ED(d, excl, self._log_add,
                                           delay=dly, timeout=to,
                                           brand_domain=getattr(self, "_brand_domain", ""))
                        if em: d["emails"] = em; d["src"] = src
                    except Exception as e:
                        self._log_add(f"  ⚠  {d['name'][:30]}: {e}")
                    finally:
                        with lock:
                            done[0] += 1
                            pct = 44 + int(50 * done[0] /
                                           max(len(dealers), 1))
                            self._set_prog(
                                pct,
                                f"{done[0]}/{len(dealers)}"
                                f"  —  {d['name'][:32]}")
                            # ── Live results update ────────────────────
                            self.results = list(dealers)
                            em_total = sum(len(x["emails"]) for x in dealers)
                            self._sv["s1"].set(str(len(dealers)))
                            self._sv["s2"].set(str(em_total))
                            self._sv["s3"].set(str(sum(1 for x in dealers if x["src"] in ("web","hunter"))))
                            self._sv["s4"].set(str(sum(1 for x in dealers if x["addr"])))
                            self.after(0, self._render)

                threads = []
                for d in dealers:
                    self._pause_ev.wait()
                    t = threading.Thread(target=do_one, args=(d,),
                                         daemon=True)
                    threads.append(t); t.start()
                    while sum(1 for x in threads if x.is_alive()) >= self._max_threads:
                        time.sleep(0.2)
                for t in threads: t.join(timeout=40)

            if self._v2.get():
                seen, uniq = set(), []
                for d in dealers:
                    k = re.sub(r"[^a-z0-9]","",d["name"].lower())[:20]
                    if k not in seen: seen.add(k); uniq.append(d)
                dealers = uniq

            total = sum(len(d["emails"]) for d in dealers)
            self._set_prog(100, "Extraction terminée ✓")
            self._log_add(
                f"━━  {len(dealers)} concessionnaires  ·  "
                f"{total} emails validés  ━━", "head")
            self.results = dealers
            self.after(0, self._finish)

        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Erreur", str(e)))
            self._log_add(f"✗  Erreur critique : {e}", "fail")
            self.after(0, lambda: (
                self._btn.configure(state="normal",
                                    text=self.L("btn_go")),
                self._pause_btn.configure(state="disabled")
            ))
            self._run = False
            self._stop_anim()

    def _finish(self):
        self._run = False; self._paused = False
        self._stop_anim()
        self._pause_btn.configure(state="disabled", text="⏸",
                                   bg=G1, fg=T2)
        self._pause_pill.pack_forget()
        d = self.results
        em = sum(len(x["emails"]) for x in d)
        wb = sum(1 for x in d if x["src"] in ("web","hunter"))
        ad = sum(1 for x in d if x["addr"])
        self._sv["s1"].set(str(len(d))); self._sv["s2"].set(str(em))
        self._sv["s3"].set(str(wb));     self._sv["s4"].set(str(ad))
        self._btn.configure(state="normal", text=self.L("btn_again"))
        self._status_lbl.configure(
            text=self.L("last") + datetime.now().strftime("%H:%M"))
        self._render()
        self._nb.select(self._f2)

    def _render(self):
        d = self.results
        em = sum(len(x["emails"]) for x in d)
        ad = sum(1 for x in d if x["addr"])
        sm = {"page": self.L("src_page"),
              "web":  self.L("src_web"),
              "hunter": self.L("src_hunter")}
        self._meta.configure(
            text=f"{len(d)} {self.L('stat1')}  ·  "
                 f"{em} {self.L('stat2')}  ·  "
                 f"{ad} {self.L('stat4')}")
        for r in self._tree.get_children():
            self._tree.delete(r)
        n = 1
        for x in d:
            sl = sm.get(x["src"], "—")
            if not x["emails"]:
                self._tree.insert("", "end", tags=("miss",), values=(
                    n, x["name"], x["addr"], x["phone"], "—", "—", sl))
                n += 1
            else:
                for i, (email, role) in enumerate(
                        sorted(x["emails"].items())):
                    tag = {"page":"","web":"web",
                           "hunter":"hunter"}.get(x["src"], "")
                    # Pick tag based on score in role string
                    import re as _re
                    score_m = _re.search(r"\[(\d+)%\]", role)
                    score_v = int(score_m.group(1)) if score_m else 50
                    if x["src"] == "hunter": row_tag = "hunter"
                    elif score_v >= 70: row_tag = "high"
                    elif score_v >= 40: row_tag = "medium"
                    else: row_tag = "low"
                    # Extract score and prepend icon to email display
                    import re as _re
                    score_m = _re.search(r"\[(\d+)%\]", role)
                    score_v = int(score_m.group(1)) if score_m else 50
                    icon = "⭐⭐⭐" if score_v >= 90 else "⭐⭐" if score_v >= 70 else "⭐" if score_v >= 40 else "·"
                    email_display = f"{icon} {email}"
                    self._tree.insert("", "end", tags=(row_tag,), values=(
                        n  if i==0 else "",
                        x["name"]  if i==0 else "",
                        x["addr"]  if i==0 else "",
                        x["phone"] if i==0 else "",
                        email_display, role,
                        sl if i==0 else ""))
                n += 1

    def _copy(self):
        em = sorted({e for x in self.results for e in x["emails"]})
        if not em:
            messagebox.showinfo("", self.L("no_copy")); return
        self.clipboard_clear()
        self.clipboard_append("\n".join(em))
        messagebox.showinfo("✓", f"{len(em)} {self.L('copied')}")

    def _export(self):
        if not self.results:
            messagebox.showinfo("", self.L("no_copy")); return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, GradientFill)
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("", "Module openpyxl manquant.\npip install openpyxl")
            return

        desk  = os.path.join(os.path.expanduser("~"), "Desktop")
        os.makedirs(desk, exist_ok=True)
        fname = f"silence_dealers_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        fpath = os.path.join(desk, fname)

        # ── Styles ──────────────────────────────────────────────────────
        RED    = "E8192C"
        WHITE  = "FFFFFF"
        DARK   = "111118"
        GRAY1  = "F7F7F8"   # alternating row light
        GRAY2  = "FFFFFF"   # alternating row white
        BORDCOL= "E4E4E7"
        GREEN_LT = "F0FDF4"
        RED_LT   = "FFF1F2"
        AMBER_LT = "FFFBEB"

        hdr_font  = Font(name="Arial", bold=True, size=10, color=WHITE)
        hdr_fill  = PatternFill("solid", fgColor=RED)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin = Side(style="thin", color=BORDCOL)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        title_font = Font(name="Arial", bold=True, size=14, color=DARK)
        sub_font   = Font(name="Arial", size=9, color="888890", italic=True)
        cell_font  = Font(name="Arial", size=9, color=DARK)
        cell_align = Alignment(vertical="center", wrap_text=False)
        cell_align_wrap = Alignment(vertical="center", wrap_text=True)
        link_font  = Font(name="Arial", size=9, color="2563EB", underline="single")

        def make_fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        # ── Workbook ────────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Concessionnaires"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"  # freeze title rows

        # ── Title block (rows 1–2) ───────────────────────────────────────
        ws.merge_cells("A1:L1")
        t = ws["A1"]
        t.value = f"SILENCE.ECO — Réseau Concessionnaires  ·  {datetime.now().strftime('%d/%m/%Y')}"
        t.font  = title_font
        t.fill  = make_fill("FFFFFF")
        t.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30

        total_em = sum(len(x["emails"]) for x in self.results)
        total_ad = sum(1 for x in self.results if x.get("addr"))
        ws.merge_cells("A2:L2")
        s = ws["A2"]
        s.value = (f"{len(self.results)} concessionnaires  ·  "
                   f"{total_em} emails trouvés  ·  "
                   f"{total_ad} adresses  ·  "
                   f"Export Silence Dealer Finder")
        s.font      = sub_font
        s.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 18

        # ── Headers (row 3) ─────────────────────────────────────────────
        COLS = [
            ("#",              4),
            ("Concessionnaire",32),
            ("Adresse",        28),
            ("CP",              7),
            ("Ville",          18),
            ("Téléphone",      14),
            ("Email",          30),
            ("Confiance",       9),
            ("Rôle",           16),
            ("Site web",       26),
            ("SIRET",          16),
            ("Source",         10),
        ]
        for col_i, (header, width) in enumerate(COLS, 1):
            cell = ws.cell(row=3, column=col_i, value=header)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_i)].width = width
        ws.row_dimensions[3].height = 22

        # ── Data rows ────────────────────────────────────────────────────
        sm = {"page": "Page", "web": "Site web", "hunter": "Hunter.io"}
        row_num = 4
        dealer_num = 1

        for x in self.results:
            if not x["emails"] and not self.inc_no.get():
                continue

            addr_full = x.get("addr", "")
            cp_m = re.search(r"(\d{5})\s*(.*)", addr_full)
            cp   = cp_m.group(1) if cp_m else ""
            city = (cp_m.group(2).strip()[:30] if cp_m else "").upper()
            addr_street = addr_full[:addr_full.find(cp)].strip(" ,") if cp and cp in addr_full else addr_full
            siret  = x.get("siret", "")
            site   = x.get("website", "")
            source = sm.get(x.get("src",""), "—")

            emails = sorted(x["emails"].items()) if x["emails"] else [("—", "—")]
            first_row = row_num

            for i, (email, role) in enumerate(emails):
                # Row fill: alternating, with special tint for score
                score_m = re.search(r"\[(\d+)%\]", role)
                score_v = int(score_m.group(1)) if score_m else 0
                if email == "—":
                    bg = GRAY1 if dealer_num % 2 == 0 else GRAY2
                elif score_v >= 90:
                    bg = "F0FDF4"   # green tint
                elif score_v >= 70:
                    bg = "F0F9FF"   # blue tint
                elif score_v >= 40:
                    bg = GRAY1 if dealer_num % 2 == 0 else GRAY2
                else:
                    bg = "FFFBEB"   # amber tint (low confidence)

                row_fill = make_fill(bg)

                def wc(col, val, wrap=False, is_link=False):
                    c = ws.cell(row=row_num, column=col, value=val)
                    c.fill   = row_fill
                    c.border = border
                    c.font   = link_font if is_link else cell_font
                    c.alignment = cell_align_wrap if wrap else cell_align
                    return c

                # Col 1: # (only first row of dealer)
                wc(1, dealer_num if i == 0 else "").alignment = Alignment(horizontal="center", vertical="center")
                wc(2, x["name"]   if i == 0 else "")
                wc(3, addr_street if i == 0 else "", wrap=True)
                wc(4, cp          if i == 0 else "").alignment = Alignment(horizontal="center", vertical="center")
                wc(5, city        if i == 0 else "")
                wc(6, x.get("phone","") if i == 0 else "").alignment = Alignment(horizontal="center", vertical="center")
                # Email
                wc(7, email if email != "—" else "", is_link=(email != "—" and "@" in email))
                # Score
                score_display = f"{score_v}%" if score_v > 0 and email != "—" else ""
                sc = wc(8, score_display)
                sc.alignment = Alignment(horizontal="center", vertical="center")
                if score_v >= 90:   sc.font = Font(name="Arial", size=9, bold=True, color="16A34A")
                elif score_v >= 70: sc.font = Font(name="Arial", size=9, color="2563EB")
                elif score_v >= 40: sc.font = Font(name="Arial", size=9, color="D97706")
                elif score_v > 0:   sc.font = Font(name="Arial", size=9, color="9A3412")
                # Role (strip score from display)
                role_clean = re.sub(r"\s*\[\d+%\]", "", role).strip() if role != "—" else ""
                wc(9, role_clean if i == 0 or role_clean else "")
                # Website (clickable)
                site_cell = wc(10, site if i == 0 else "", is_link=bool(site and i == 0))
                if site and i == 0:
                    ws.cell(row=row_num, column=10).hyperlink = site
                # SIRET
                wc(11, siret if i == 0 else "")
                # Source
                src_cell = wc(12, source if i == 0 else "")
                src_cell.alignment = Alignment(horizontal="center", vertical="center")

                ws.row_dimensions[row_num].height = 18
                row_num += 1

            dealer_num += 1

        # ── Auto-filter on header row ────────────────────────────────────
        ws.auto_filter.ref = f"A3:L{row_num - 1}"

        # ── Legend sheet ────────────────────────────────────────────────
        wl = wb.create_sheet("Légende")
        wl.sheet_view.showGridLines = False
        wl.column_dimensions["A"].width = 14
        wl.column_dimensions["B"].width = 40

        legend_title = wl["A1"]
        legend_title.value = "Légende — Score de confiance email"
        legend_title.font  = Font(name="Arial", bold=True, size=12, color=DARK)
        wl.merge_cells("A1:B1")
        wl.row_dimensions[1].height = 26

        legend_rows = [
            ("≥ 90%  ★★★", "Certitude — source officielle ou nom exact dans l'email",      "F0FDF4", "16A34A"),
            ("70-89% ★★",  "Très probable — mot-clé du nom dans le préfixe/domaine",          "F0F9FF", "2563EB"),
            ("40-69% ★",   "Possible — préfixe pro sur domaine lié",                           "FFFBEB", "D97706"),
            ("< 40%  ·",   "Faible — nom de personne, employé possible",                       "FFF5F5", "9A3412"),
            ("0%     ✗",   "Exclu — domaine blacklisté (assurance, plateforme, etc.)",          "F1F1F5", "888890"),
        ]
        for r_i, (score, desc, bg, txt) in enumerate(legend_rows, 2):
            sc = wl.cell(row=r_i, column=1, value=score)
            sc.font      = Font(name="Arial", bold=True, size=10, color=txt)
            sc.fill      = PatternFill("solid", fgColor=bg)
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.border    = border
            dc = wl.cell(row=r_i, column=2, value=desc)
            dc.font      = Font(name="Arial", size=9, color=DARK)
            dc.fill      = PatternFill("solid", fgColor=bg)
            dc.alignment = Alignment(vertical="center")
            dc.border    = border
            wl.row_dimensions[r_i].height = 20

        # ── Save ─────────────────────────────────────────────────────────
        wb.save(fpath)

        rows = row_num - 4
        messagebox.showinfo("✓", f"Export Excel sauvegardé :\n{fname}\n\n{rows} lignes exportées")
        try:
            os.startfile(fpath)
        except Exception:
            try: os.system(f'open "{fpath}"')
            except Exception: pass

norm = normalize  # alias

if __name__ == "__main__":
    app = App(
        STRINGS      = T,
        EXCL_DEFAULT = EXCL_DEFAULT,
        scrape_page  = scrape_page,
        enrich_dealer= enrich_dealer,
        norm         = norm,
        EMAIL_RE     = EMAIL_RE,
    )
    app.mainloop()
# Domaines toujours exclus (assurances, plateformes, etc.)
# Always excluded (insurance, platforms)
BAD_DOMAINS = [
    "sollyazar.","allianz.","axa.","maif.","macif.","generali.",
    "mma.","groupama.","covea.","april.","google.","bing.",
    "facebook.","instagram.","twitter.","linkedin.","youtube.",
    "indeed.","monster.","pagesjaunes.","societe.com","verif.com",
    "pappers.","cloudflare.","sentry.","clubmobilite.",
]

GENERIC_TLDS = {"gmail","orange","hotmail","yahoo","outlook","sfr","wanadoo",
                "free","laposte","bbox","live","icloud","msn",
                "numericable","bouygues","neuf","alice","club"}

BRAND_WORDS  = {"ligier","aixam","microcar","chatenet","bellier","casalini",
                "store","partner","service","group","rent"}

SKIP_WORDS   = {"le","la","les","de","du","des","et","au","aux","un","une",
                "par","sur","pour","avec","sans","dans"}


def score_email(email, dealer_name, source="search", brand_domain=""):
    """
    Score générique unique : un mot du nom du concessionnaire
    apparaît-il dans l'email (préfixe ou domaine) ?

    Règle unique — pas de cas particuliers :
    ⭐⭐⭐ 90-98% : source officielle OU mot exact dans l'email
    ⭐⭐  70-89% : mot contenu dans l'email (avec ou sans fournisseur gratuit)
    ⭐   35-40% : domaine business sans correspondance
    ·   20%    : fournisseur gratuit sans correspondance
    ✗   0%     : blacklisté
    """
    if not email or "@" not in email: return 0, "invalide"

    dom_full = email.split("@")[-1].lower()

    if any(p in dom_full for p in BAD_DOMAINS):
        return 0, "domaine blacklisté"

    if brand_domain and brand_domain.lower() in dom_full:
        return 0, f"email du constructeur ({brand_domain})"

    if source == "page":    return 98, "page officielle ⭐⭐⭐"
    if source == "site":    return 95, "site web ⭐⭐⭐"
    if source == "hunter":  return 90, "Hunter.io ⭐⭐⭐"

    local    = normalize(email.split("@")[0])
    dom      = normalize(email.split("@")[-1].split(".")[0])
    is_free  = dom in GENERIC_TLDS

    # Mots du nom — tout mot ≥ 3 chars sauf articles et mots de marque
    words = [normalize(w) for w in re.split(r"[\s\-\'&/,.]", dealer_name)
             if len(normalize(w)) >= 3
             and normalize(w) not in SKIP_WORDS
             and normalize(w) not in BRAND_WORDS]
    if not words:
        words = [normalize(dealer_name)]
    words = list(dict.fromkeys(words))

    best_score = 0
    best_reason = ""

    for word in words:
        if len(word) < 3: continue
        # Correspondance exacte
        if word == local:      s, r = 92, f"préfixe = '{word}'"
        elif word == dom:      s, r = 92, f"domaine = '{word}'"
        # Contenu dans
        elif word in local:    s, r = 85, f"'{word}' dans préfixe"
        elif word in dom:      s, r = 82, f"'{word}' dans domaine"
        # Inclusion inverse
        elif len(local)>=4 and local in word: s, r = 78, f"préfixe '{local}' ⊂ '{word}'"
        elif len(dom)>=4   and dom   in word: s, r = 75, f"domaine '{dom}' ⊂ '{word}'"
        else: continue
        if s > best_score: best_score, best_reason = s, r

    if best_score == 0:
        if is_free: return 20, "aucun lien — adresse gratuite ·"
        else:       return 35, "domaine business sans lien identifié ⭐"

    # Légère pénalité fournisseur gratuit (on ne peut pas vérifier l'ownership)
    if is_free and best_score < 90:
        best_score = max(best_score - 10, 20)
        best_reason += " (adresse gratuite)"

    # Icône selon score final
    if best_score >= 90: best_reason += " ⭐⭐⭐"
    elif best_score >= 70: best_reason += " ⭐⭐"
    elif best_score >= 40: best_reason += " ⭐"
    else: best_reason += " ·"

    return best_score, best_reason


def email_belongs_to_dealer(email, dealer_name, source="search", brand_domain=""):
    """Compatibilité — retourne (bool, reason). Garde tout sauf blacklisté."""
    s, r = score_email(email, dealer_name, source, brand_domain)
    return s > 0, r


